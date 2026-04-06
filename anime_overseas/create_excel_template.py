"""
动漫出海 - 运营数据 Excel 模板生成器
anime_overseas/create_excel_template.py

生成以下 Excel 表格:
1. 每日运营日志
2. 内容发布记录
3. 收益追踪表
4. 授权管理表

用法:
    python create_excel_template.py
"""

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

from datetime import datetime, date, timedelta

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


def create_styles():
    """创建样式字典"""
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    return {
        "header_dark": {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill("solid", fgColor="1a1a2e"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": thin_border,
        },
        "header_blue": {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill("solid", fgColor="0f3460"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": thin_border,
        },
        "header_green": {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill("solid", fgColor="28a745"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": thin_border,
        },
        "cell_normal": {
            "font": Font(size=10),
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": thin_border,
        },
        "cell_center": {
            "font": Font(size=10),
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": thin_border,
        },
        "cell_money": {
            "font": Font(size=10, color="28a745", bold=True),
            "fill": PatternFill("solid", fgColor="f0fff4"),
            "alignment": Alignment(horizontal="right", vertical="center"),
            "border": thin_border,
        },
        "cell_alert": {
            "font": Font(size=10, color="dc3545", bold=True),
            "fill": PatternFill("solid", fgColor="fff8f8"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": thin_border,
        },
        "cell_highlight": {
            "font": Font(size=10, bold=True),
            "fill": PatternFill("solid", fgColor="fff9e6"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": thin_border,
        },
        "title": {
            "font": Font(bold=True, size=16, color="1a1a2e"),
            "fill": PatternFill("solid", fgColor="f8f9fa"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
    }


def apply_style(cell, style_dict):
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def create_operations_log(wb):
    """创建每日运营日志表"""
    ws = wb.active
    ws.title = "每日运营日志"
    styles = create_styles()

    # 标题行
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = "动漫出海运营 · 每日运营日志"
    apply_style(title_cell, styles["title"])
    ws.row_dimensions[1].height = 30

    # 列标题
    headers = [
        "日期", "平台", "发布数", "播放量", "新增订阅", "总订阅",
        "广告收益", "带货收益", "商单收益", "总收益(净)",
        "爆款数(>1万)", "版权投诉", "备注"
    ]
    ws.append([""] + headers)
    for col, header in enumerate(headers, 2):
        apply_style(ws.cell(2, col), styles["header_dark"])
    ws.row_dimensions[2].height = 25

    # 数据验证（下拉列表）
    platform_dv = DataValidation(
        type="list", formula1='"YouTube,TikTok,Instagram,Facebook,All"',
        allow_blank=True
    )
    ws.add_data_validation(platform_dv)
    platform_dv.add(f"B3:B100")

    # 示例数据
    today = date.today()
    for i in range(30):
        row_date = today - timedelta(days=29-i)
        row_num = i + 3
        row = [
            row_date.strftime("%Y-%m-%d"),
            "All",
            "", "", "", "",
            "", "", "", "",
            "", "", ""
        ]
        ws.append([""] + row)
        for col in range(2, 14):
            style = styles["cell_center"] if col in [2, 4, 5, 6, 11] else styles["cell_normal"]
            apply_style(ws.cell(row_num, col), style)

    # 列宽
    col_widths = [3, 12, 8, 10, 10, 10, 10, 10, 10, 10, 10, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结窗格
    ws.freeze_panes = "C3"

    return ws


def create_content_log(wb):
    """创建内容发布记录表"""
    ws = wb.create_sheet("内容发布记录")
    styles = create_styles()

    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "动漫出海运营 · 内容发布记录"
    apply_style(title_cell, styles["title"])
    ws.row_dimensions[1].height = 30

    headers = [
        "日期", "平台", "IP名称", "集数", "内容类型",
        "标题", "标签", "时长(s)", "播放量",
        "点赞数", "收益(净)", "爆款?"
    ]
    ws.append([""] + headers)
    for col, header in enumerate(headers, 2):
        apply_style(ws.cell(2, col), styles["header_blue"])
    ws.row_dimensions[2].height = 25

    # 下拉验证
    dv_platform = DataValidation(type="list", formula1='"YouTube,TikTok,Instagram,Facebook,Pinterest"', allow_blank=True)
    dv_type = DataValidation(type="list", formula1='"clip,analysis,review,mashup,static,fanwork"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_platform)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_yesno)
    dv_platform.add("B3:B500")
    dv_type.add("E3:E500")
    dv_yesno.add("L3:L500")

    # 示例行
    for i in range(50):
        row_num = i + 3
        row = [""] + [""] * 11
        ws.append(row)
        for col in range(2, 13):
            style = styles["cell_center"] if col in [2, 4, 8, 9, 10, 11] else styles["cell_normal"]
            apply_style(ws.cell(row_num, col), style)

    col_widths = [3, 12, 15, 8, 12, 35, 30, 8, 10, 10, 10, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "C3"
    return ws


def create_revenue_tracking(wb):
    """创建收益追踪表"""
    ws = wb.create_sheet("收益追踪")
    styles = create_styles()

    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "动漫出海运营 · 月度收益追踪"
    apply_style(title_cell, styles["title"])
    ws.row_dimensions[1].height = 30

    headers = [
        "日期", "平台", "收益类型", "内容/IP",
        "毛收益($)", "平台扣减(%)", "净收益($)",
        "地区", "IP授权方", "备注"
    ]
    ws.append([""] + headers)
    for col, header in enumerate(headers, 2):
        apply_style(ws.cell(2, col), styles["header_green"])
    ws.row_dimensions[2].height = 25

    dv_platform = DataValidation(type="list", formula1='"youtube,tiktok,instagram,amazon,temu,gumroad,commission,order,other"', allow_blank=True)
    dv_type = DataValidation(type="list", formula1='"广告分成,商单,带货佣金,版权代理,二创服务,会员分销,其他"', allow_blank=True)
    ws.add_data_validation(dv_platform)
    ws.add_data_validation(dv_type)
    dv_platform.add("B3:B1000")
    dv_type.add("C3:C1000")

    # 示例数据
    for i in range(100):
        row_num = i + 3
        row = [""] + [""] * 9
        ws.append(row)
        for col in range(2, 11):
            apply_style(ws.cell(row_num, col), styles["cell_normal"])
        apply_style(ws.cell(row_num, 6), styles["cell_money"])

    col_widths = [3, 12, 12, 20, 12, 12, 12, 10, 15, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "C3"
    return ws


def create_ip_management(wb):
    """创建IP授权管理表"""
    ws = wb.create_sheet("IP授权管理")
    styles = create_styles()

    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "动漫出海运营 · IP授权管理"
    apply_style(title_cell, styles["title"])
    ws.row_dimensions[1].height = 30

    headers = [
        "IP名称", "版权方", "授权类型", "授权地区",
        "起始日期", "结束日期", "佣金比例(%)",
        "月费/分成($)", "状态", "联系人"
    ]
    ws.append([""] + headers)
    for col, header in enumerate(headers, 2):
        apply_style(ws.cell(2, col), styles["header_dark"])
    ws.row_dimensions[2].height = 25

    dv_auth = DataValidation(type="list", formula1='"非独家,独家,试运行,已过期,洽谈中"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"有效,待签署,洽谈中,已过期,不续约"', allow_blank=True)
    ws.add_data_validation(dv_auth)
    ws.add_data_validation(dv_status)
    dv_auth.add("C3:C200")
    dv_status.add("I3:I200")

    # 预填一些示例IP
    sample_ips = [
        ["斗罗大陆", "腾讯动漫", "非独家", "东南亚", "", "", "25", "", "有效", "overseas@tencent.com"],
        ["凡人修仙传", "起点中文网", "非独家", "全球", "", "", "20", "", "洽谈中", "international@qidian.com"],
        ["狐妖小红娘", "腾讯动漫", "非独家", "东南亚/中东", "", "", "25", "", "有效", "overseas@tencent.com"],
        ["一人之下", "腾讯动漫", "非独家", "东南亚", "", "", "25", "", "洽谈中", "overseas@tencent.com"],
        ["快看漫画合集", "快看漫画", "非独家", "全球", "", "", "30", "", "洽谈中", "global@kuaikanwc.com"],
    ]

    for i, ip in enumerate(sample_ips):
        row_num = i + 3
        ws.append([""] + ip)
        for col in range(2, 11):
            style = styles["cell_highlight"] if col == 9 else styles["cell_normal"]
            apply_style(ws.cell(row_num, col), style)

    # 空行
    for i in range(20):
        row_num = i + 3 + len(sample_ips)
        ws.append([""] + [""] * 9)
        for col in range(2, 11):
            apply_style(ws.cell(row_num, col), styles["cell_normal"])

    col_widths = [3, 15, 18, 12, 12, 12, 12, 12, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "C3"
    return ws


def create_dashboard(wb):
    """创建数据看板总览"""
    ws = wb.create_sheet("数据看板", 0)  # 放在第一位
    styles = create_styles()

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "动漫出海运营 · 数据看板"
    apply_style(title_cell, styles["title"])
    ws.row_dimensions[1].height = 35

    # KPI 卡片
    kpi_data = [
        ("B3", "本月总播放", "=SUM('每日运营日志'!D:D)", "次", "1a1a2e"),
        ("B4", "本月总收益(净)", "=SUM('每日运营日志'!J:J)", "$", "28a745"),
        ("B5", "当前订阅总数", "=SUM('每日运营日志'!F:F)", "人", "0f3460"),
        ("B6", "本月爆款数", "=SUM('每日运营日志'!K:K)", "个", "e94560"),
    ]

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"📊 数据概览 | {datetime.now().strftime('%Y年%m月%d日 %H:%M 更新')}"
    ws["A2"].font = Font(size=11, color="666666", italic=True)
    ws["A2"].alignment = Alignment(horizontal="right")

    for cell_ref, label, formula, unit, color in kpi_data:
        ws.merge_cells(f"{cell_ref}:{cell_ref[0]}{int(cell_ref[1])+1}")
        label_cell = ws[cell_ref]
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=color)
        label_cell.font = Font(bold=True, color="FFFFFF", size=12)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium")
        )
        ws.row_dimensions[int(cell_ref[1])].height = 40

        ws.merge_cells(f"{get_column_letter(3)}{int(cell_ref[1])}:{get_column_letter(4)}{int(cell_ref[1])}")
        val_cell = ws[f"C{int(cell_ref[1])}"]
        val_cell.value = f"={formula}"
        val_cell.number_format = '#,##0' if unit == "次" else '"$"#,##0.00'
        val_cell.font = Font(bold=True, size=18, color=color)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 平台分布表
    row_start = 10
    ws.merge_cells(f"A{row_start}:H{row_start}")
    ws[f"A{row_start}"].value = "📺 平台收益分布"
    ws[f"A{row_start}"].font = Font(bold=True, size=13, color="1a1a2e")
    ws.row_dimensions[row_start].height = 25

    platform_headers = ["平台", "本月播放", "本月收益", "占总收益%", "环比上周%", "状态"]
    for col, header in enumerate(platform_headers, 2):
        cell = ws.cell(row_start + 1, col)
        cell.value = header
        apply_style(cell, styles["header_dark"])

    platforms = ["YouTube", "TikTok", "Instagram", "Facebook", "联盟带货", "版权代理"]
    for i, platform in enumerate(platforms):
        row = row_start + 2 + i
        ws.cell(row, 2).value = platform
        for col in range(2, 8):
            apply_style(ws.cell(row, col), styles["cell_normal"])
        ws.cell(row, 3).number_format = "#,##0"
        ws.cell(row, 4).number_format = '"$"#,##0.00'
        ws.cell(row, 5).number_format = "0.0%"
        ws.cell(row, 6).number_format = "0.0%"
        ws.cell(row, 7).value = "📈 正常"

    # 内容类型分析
    row_start2 = 20
    ws.merge_cells(f"A{row_start2}:H{row_start2}")
    ws[f"A{row_start2}"].value = "🎬 内容类型分析"
    ws[f"A{row_start2}"].font = Font(bold=True, size=13, color="1a1a2e")
    ws.row_dimensions[row_start2].height = 25

    type_headers = ["内容类型", "发布数", "平均播放", "最高播放", "总收益", "占比"]
    for col, header in enumerate(type_headers, 2):
        cell = ws.cell(row_start2 + 1, col)
        cell.value = header
        apply_style(cell, styles["header_blue"])

    content_types = ["精彩剪辑", "动漫解说", "TOP盘点", "壁纸推荐", "二创混剪", "资讯类"]
    for i, ctype in enumerate(content_types):
        row = row_start2 + 2 + i
        ws.cell(row, 2).value = ctype
        for col in range(2, 8):
            apply_style(ws.cell(row, col), styles["cell_normal"])

    # 列宽
    ws.column_dimensions["A"].width = 3
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # 冻结
    ws.freeze_panes = "A3"
    return ws


def create_excel_template(output_path: str = None):
    """生成完整的 Excel 运营模板"""
    if output_path is None:
        output_path = OUTPUT_DIR / f"动漫出海运营模板_{datetime.now().strftime('%Y%m%d')}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除默认sheet

    print("正在创建 Excel 模板...")
    print("  [1/5] 数据看板...")
    create_dashboard(wb)
    print("  [2/5] 每日运营日志...")
    create_operations_log(wb)
    print("  [3/5] 内容发布记录...")
    create_content_log(wb)
    print("  [4/5] 收益追踪...")
    create_revenue_tracking(wb)
    print("  [5/5] IP授权管理...")
    create_ip_management(wb)

    wb.save(str(output_path))
    print(f"\n✅ Excel 模板已生成: {output_path}")
    print(f"\n包含工作表:")
    for i, sheet in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet}")
    return str(output_path)


if __name__ == "__main__":
    output = create_excel_template()
    print(f"\n打开文件: {output}")
