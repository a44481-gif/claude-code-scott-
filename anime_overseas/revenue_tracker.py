"""
动漫出海 - 收益追踪与结算系统
anime_overseas/revenue_tracker.py

功能：
1. 多平台收益记录（YouTube/TikTok/联盟佣金）
2. 各地区、各内容类型收益分析
3. 自动计算净利润（税后/扣除平台分成后）
4. 发票和结算报告生成
5. 收益目标追踪（vs 上月/上月同期）
6. 财务数据导出 Excel

用法:
    tracker = RevenueTracker()
    tracker.add_revenue("youtube", 125.50, "广告分成", "斗罗大陆 EP5")
    tracker.get_monthly_report()
    tracker.export_excel("revenue_2026_04.xlsx")
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Optional

from anime_ops_config import MONETIZATION, AFFILIATE_PLATFORMS, LOG_DIR, OUTPUT_DIR

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "revenue.log", encoding="utf-8"),
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger("RevenueTracker")


# ============ 收益条目 ============

class RevenueEntry:
    def __init__(
        self,
        platform: str,
        amount: float,
        category: str,
        content_name: str = None,
        currency: str = "USD",
        date_str: str = None,
        region: str = None,
        anime_ip: str = None,
        notes: str = None,
    ):
        self.id = f"{datetime.now().timestamp()}_{hash((platform, amount, content_name)) % 100000}"
        self.platform = platform  # youtube / tiktok / affiliate / commission / order
        self.amount = float(amount)
        self.gross_amount = float(amount)
        self.category = category  # 广告分成 / 商单 / 带货佣金 / 版权代理 / 二创服务
        self.content_name = content_name
        self.currency = currency
        self.date = date_str or datetime.now().strftime("%Y-%m-%d")
        self.region = region
        self.anime_ip = anime_ip  # 关联的IP名称
        self.notes = notes
        self.net_amount = self._calculate_net()

    def _calculate_net(self) -> float:
        """计算净收益（扣除平台分成后）"""
        deductions = {
            "youtube": 0.45,      # YouTube拿走45%
            "tiktok": 0.50,       # TikTok拿走50%
            "amazon": 0.10,       # Amazon联盟10%管理费
            "temu": 0.05,        # Temu 5%
            "gumroad": 0.10,     # Gumroad 10%
            "commission": 0.25,  # 版权代理佣金（代理人自留）
        }
        rate = deductions.get(self.platform, 0)
        return round(self.gross_amount * (1 - rate), 2)

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "platform": self.platform,
            "gross_amount": self.gross_amount,
            "net_amount": self.net_amount,
            "category": self.category,
            "content_name": self.content_name,
            "currency": self.currency,
            "date": self.date,
            "region": self.region,
            "anime_ip": self.anime_ip,
            "notes": self.notes,
        }


# ============ 收益追踪器 ============

class RevenueTracker:
    """
    动漫出海收益追踪系统
    """

    def __init__(self, data_dir: str = None):
        self.data_dir = Path(data_dir or OUTPUT_DIR)
        self.data_dir.mkdir(exist_ok=True)
        self.revenue_file = self.data_dir / "revenue_records.json"
        self.entries: list[RevenueEntry] = []
        self._load()

    def _load(self):
        if self.revenue_file.exists():
            try:
                data = json.loads(self.revenue_file.read_text(encoding="utf-8"))
                for e in data.get("entries", []):
                    entry = RevenueEntry(**e)
                    entry.net_amount = e.get("net_amount", entry.net_amount)
                    self.entries.append(entry)
                logger.info(f"已加载 {len(self.entries)} 条收益记录")
            except Exception as e:
                logger.error(f"加载收益数据失败: {e}")

    def _save(self):
        data = {
            "entries": [e.to_dict() for e in self.entries],
            "last_updated": datetime.now().isoformat(),
        }
        self.revenue_file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def add_revenue(
        self,
        platform: str,
        amount: float,
        category: str,
        content_name: str = None,
        currency: str = "USD",
        date_str: str = None,
        region: str = None,
        anime_ip: str = None,
        notes: str = None,
    ) -> RevenueEntry:
        """添加收益条目"""
        entry = RevenueEntry(
            platform=platform, amount=amount, category=category,
            content_name=content_name, currency=currency,
            date_str=date_str, region=region, anime_ip=anime_ip, notes=notes
        )
        self.entries.append(entry)
        self._save()
        logger.info(f"收益已记录: {platform} +${entry.net_amount} ({category})")
        return entry

    def get_entries_by_date(self, start_date: str, end_date: str = None) -> list[RevenueEntry]:
        """获取日期范围内的收益"""
        if end_date is None:
            end_date = start_date
        return [e for e in self.entries if start_date <= e.date <= end_date]

    def get_monthly_summary(self, year: int = None, month: int = None) -> dict:
        """月度收益汇总"""
        if year is None:
            year = datetime.now().year
        if month is None:
            month = datetime.now().month

        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year+1}-01-01"
        else:
            end_date = f"{year}-{month+1:02d}-01"

        entries = self.get_entries_by_date(start_date, end_date)

        summary = {
            "year": year,
            "month": month,
            "total_gross": sum(e.gross_amount for e in entries),
            "total_net": sum(e.net_amount for e in entries),
            "by_platform": {},
            "by_category": {},
            "by_ip": {},
            "by_region": {},
            "entry_count": len(entries),
        }

        for e in entries:
            summary["by_platform"][e.platform] = summary["by_platform"].get(e.platform, 0) + e.net_amount
            summary["by_category"][e.category] = summary["by_category"].get(e.category, 0) + e.net_amount
            if e.anime_ip:
                summary["by_ip"][e.anime_ip] = summary["by_ip"].get(e.anime_ip, 0) + e.net_amount
            if e.region:
                summary["by_region"][e.region] = summary["by_region"].get(e.region, 0) + e.net_amount

        return summary

    def get_weekly_comparison(self) -> dict:
        """本周 vs 上周对比"""
        today = datetime.now().date()
        this_week_start = today - timedelta(days=today.weekday())
        last_week_start = this_week_start - timedelta(days=7)
        last_week_end = this_week_start - timedelta(days=1)

        this_week_entries = self.get_entries_by_date(
            this_week_start.strftime("%Y-%m-%d"),
            today.strftime("%Y-%m-%d")
        )
        last_week_entries = self.get_entries_by_date(
            last_week_start.strftime("%Y-%m-%d"),
            last_week_end.strftime("%Y-%m-%d")
        )

        this_week_net = sum(e.net_amount for e in this_week_entries)
        last_week_net = sum(e.net_amount for e in last_week_entries)

        growth = 0.0
        if last_week_net > 0:
            growth = (this_week_net - last_week_net) / last_week_net * 100

        return {
            "this_week": {
                "start": this_week_start.strftime("%Y-%m-%d"),
                "end": today.strftime("%Y-%m-%d"),
                "net": this_week_net,
                "gross": sum(e.gross_amount for e in this_week_entries),
                "entries": len(this_week_entries),
            },
            "last_week": {
                "start": last_week_start.strftime("%Y-%m-%d"),
                "end": last_week_end.strftime("%Y-%m-%d"),
                "net": last_week_net,
                "gross": sum(e.gross_amount for e in last_week_entries),
                "entries": len(last_week_entries),
            },
            "growth_rate": round(growth, 1),
        }

    def estimate_full_month(self, year: int = None, month: int = None) -> dict:
        """预估当月全月收益（基于当前日均）"""
        today = datetime.now()
        if year is None:
            year = today.year
        if month is None:
            month = today.month

        # 计算当月已过去天数
        if today.year == year and today.month == month:
            days_passed = today.day
        else:
            import calendar
            days_passed = calendar.monthrange(year, month)[1]

        summary = self.get_monthly_summary(year, month)
        daily_avg = summary["total_net"] / max(days_passed, 1)

        import calendar
        days_in_month = calendar.monthrange(year, month)[1]

        return {
            **summary,
            "days_passed": days_passed,
            "days_in_month": days_in_month,
            "daily_average": round(daily_avg, 2),
            "projected_full_month": round(daily_avg * days_in_month, 2),
            "progress_percentage": round(days_passed / days_in_month * 100, 1),
        }

    def check_goal_progress(self, monthly_goal: float) -> dict:
        """检查收益目标完成进度"""
        estimate = self.estimate_full_month()
        current = estimate["total_net"]
        projected = estimate["projected_full_month"]

        return {
            "monthly_goal": monthly_goal,
            "current_net": current,
            "projected": projected,
            "progress": round(current / monthly_goal * 100, 1),
            "is_on_track": projected >= monthly_goal,
            "needed_daily": round((monthly_goal - current) / max(estimate["days_in_month"] - estimate["days_passed"], 1), 2),
        }

    def generate_monthly_report(self, year: int = None, month: int = None) -> str:
        """生成月度收益报告"""
        if year is None:
            year = datetime.now().year
        if month is None:
            month = datetime.now().month

        summary = self.get_monthly_summary(year, month)
        estimate = self.estimate_full_month(year, month)
        weekly = self.get_weekly_comparison()

        report = f"""
{'='*60}
动漫出海运营 · {year}年{month:02d}月 收益报告
{'='*60}
📅 报告日期: {datetime.now().strftime('%Y-%m-%d %H:%M')}

💰 收益概览
───────────────────────────
  总收益（毛）: ${summary['total_gross']:.2f}
  总收益（净）: ${summary['total_net']:.2f}
  记录条目: {summary['entry_count']} 条

📊 月度预估
───────────────────────────
  当月进度: {estimate['days_passed']}/{estimate['days_in_month']} 天 ({estimate['progress_percentage']}%)
  日均收益: ${estimate['daily_average']:.2f}
  预估全月: ${estimate['projected_full_month']:.2f}

📈 平台分布（净收益）
───────────────────────────
"""
        for platform, amount in sorted(summary["by_platform"].items(), key=lambda x: x[1], reverse=True):
            pct = amount / max(summary["total_net"], 1) * 100
            bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
            report += f"  {platform:12s} ${amount:8.2f} ({pct:5.1f}%) {bar}\n"

        report += f"""
📦 收益分类
───────────────────────────
"""
        for cat, amount in sorted(summary["by_category"].items(), key=lambda x: x[1], reverse=True):
            pct = amount / max(summary["total_net"], 1) * 100
            report += f"  {cat:12s} ${amount:8.2f} ({pct:5.1f}%)\n"

        if summary["by_ip"]:
            report += f"""
🎬 IP收益贡献
───────────────────────────
"""
            for ip, amount in sorted(summary["by_ip"].items(), key=lambda x: x[1], reverse=True)[:5]:
                pct = amount / max(summary["total_net"], 1) * 100
                report += f"  {ip:15s} ${amount:8.2f} ({pct:5.1f}%)\n"

        report += f"""
📆 周对比
───────────────────────────
  本周收益: ${weekly['this_week']['net']:.2f} ({weekly['this_week']['entries']}条)
  上周收益: ${weekly['last_week']['net']:.2f} ({weekly['last_week']['entries']}条)
  周环比:   {'+' if weekly['growth_rate'] >= 0 else ''}{weekly['growth_rate']:.1f}%

🎯 YPP 达标进度
───────────────────────────
"""
        for platform in ["youtube"]:
            threshold = 100 if platform == "youtube" else 0
            report += f"  YouTube: 距离 ${threshold}/月 还需 ${max(0, threshold - estimate['total_net']):.2f}\n"

        report += f"""
{'='*60}
"""
        return report

    def export_excel(self, output_path: str = None):
        """导出Excel报表"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl 未安装，请运行: pip install openpyxl")
            return None

        if output_path is None:
            output_path = str(self.data_dir / f"revenue_{datetime.now().strftime('%Y%m')}.xlsx")
        output_path = Path(output_path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "收益记录"

        # 样式
        header_fill = PatternFill("solid", fgColor="1a1a2e")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        money_font = Font(color="28a745", bold=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["日期", "平台", "收益类别", "内容/项目", "毛收益", "净收益", "货币", "地区", "IP", "备注"]
        ws.append(headers)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for entry in self.entries[-200:]:  # 最近200条
            ws.append([
                entry.date, entry.platform, entry.category, entry.content_name,
                entry.gross_amount, entry.net_amount, entry.currency,
                entry.region or "", entry.anime_ip or "", entry.notes or ""
            ])

        # 金额列格式化
        for row in ws.iter_rows(min_row=2, max_col=10):
            row[4].number_format = '"$"#,##0.00'
            row[5].number_format = '"$"#,##0.00'
            for cell in row:
                cell.border = thin_border

        # 列宽
        col_widths = [12, 10, 12, 20, 12, 12, 8, 10, 15, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 汇总Sheet
        summary = self.get_monthly_summary()
        ws2 = wb.create_sheet("月度汇总")
        ws2.append(["月份", "毛收益", "净收益", "记录数"])
        ws2.cell(1, 1).fill = header_fill
        ws2.cell(1, 1).font = header_font
        for i in range(2, 5):
            ws2.cell(1, i).fill = header_fill
            ws2.cell(1, i).font = header_font

        ws2.append([f"{summary['year']}-{summary['month']:02d}", summary['total_gross'],
                     summary['total_net'], summary['entry_count']])
        ws2.cell(2, 2).number_format = '"$"#,##0.00'
        ws2.cell(2, 3).number_format = '"$"#,##0.00'

        for col in range(1, 5):
            ws2.column_dimensions[get_column_letter(col)].width = 18

        wb.save(str(output_path))
        logger.info(f"Excel 导出: {output_path}")
        return str(output_path)


# ============ 版权代理佣金计算 ============

class CommissionCalculator:
    """版权代理佣金计算器"""

    @staticmethod
    def calculate_commission(
        gross_revenue: float,
        commission_rate: float = None,
        platform: str = "commission",
    ) -> dict:
        """计算版权代理佣金"""
        if commission_rate is None:
            commission_rate = MONETIZATION.get("版权代理佣金比例", 0.25)

        gross = gross_revenue
        agent_share = gross * commission_rate
        cp_share = gross * (1 - commission_rate)

        return {
            "gross_revenue": gross,
            "commission_rate": commission_rate,
            "agent_net_commission": round(agent_share, 2),
            "cp_share": round(cp_share, 2),
            "platform": platform,
        }


if __name__ == "__main__":
    print("=" * 60)
    print("动漫出海收益追踪系统")
    print("=" * 60)

    tracker = RevenueTracker()

    # 演示：添加一些模拟收益
    print("\n添加模拟收益数据...")
    tracker.add_revenue("youtube", 125.50, "广告分成", "斗罗大陆 EP5", region="english")
    tracker.add_revenue("youtube", 89.00, "广告分成", "凡人修仙传 EP12", region="southeast_asia")
    tracker.add_revenue("tiktok", 45.20, "Creator Fund", "斗破苍穹 EP3", region="middle_east")
    tracker.add_revenue("amazon", 32.50, "联盟带货", "手办-斗罗大陆", region="english", anime_ip="斗罗大陆")
    tracker.add_revenue("commission", 500.00, "版权代理", "凡人修仙传东南亚授权", region="southeast_asia", anime_ip="凡人修仙传")

    # 生成报告
    report = tracker.generate_monthly_report()
    print(report)

    # 检查目标进度（假设月目标$2000）
    goal = tracker.check_goal_progress(2000)
    print(f"💵 收益目标: ${goal['monthly_goal']}")
    print(f"   当前进度: {goal['progress']}%")
    print(f"   预估全月: ${goal['projected']:.2f}")
    print(f"   是否达标: {'✅ 是' if goal['is_on_track'] else '❌ 否'}")
    if not goal['is_on_track']:
        print(f"   每日需再赚: ${goal['needed_daily']:.2f}")
