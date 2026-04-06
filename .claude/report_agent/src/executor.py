"""
Report Agent - 主執行器（整合所有 Agent 數據生成報告）
"""
import json
import glob
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional
from loguru import logger

from core import MiniMaxAnalyzer, EmailSender, LarkClient


class ReportExecutor:
    """Report Agent 主執行器"""

    def __init__(self, cfg: Dict):
        self.cfg = cfg
        self.analyzer = MiniMaxAnalyzer(cfg.get("ai", {}))
        self.today_str = datetime.now().strftime("%Y%m%d")
        self.base_dir = Path(__file__).parent.parent
        self.output_dir = self.base_dir / cfg.get("output", {}).get("output_dir", "reports")
        self.output_dir.mkdir(exist_ok=True)
        self.data: Dict[str, Any] = {}

    # ── Aggregation ─────────────────────────────────────────────

    def aggregate_all(self) -> Dict[str, Any]:
        """從所有 Agent 收集數據"""
        logger.info("開始匯總所有 Agent 數據...")
        data_sources = self.cfg.get("data_sources", {})

        aggregated = {
            "it_news": {"items": [], "summary": {}, "date": ""},
            "pc_parts": {"items": [], "alerts": [], "summary": {}},
            "msi_updates": {"items": [], "new_count": 0, "summary": {}},
        }

        # IT News
        news_path = Path(data_sources.get("it_news", {}).get("path", "../it_news_agent/data"))
        news_files = sorted(news_path.glob("news_*.json"))
        if news_files:
            with open(news_files[-1], encoding="utf-8") as f:
                news_data = json.load(f)
            aggregated["it_news"]["items"] = news_data.get("items", [])
            aggregated["it_news"]["date"] = news_data.get("date", "")
            cats = {}
            for item in news_data.get("items", []):
                cat = item.get("category", "General")
                cats[cat] = cats.get(cat, 0) + 1
            aggregated["it_news"]["summary"] = {
                "total": len(news_data.get("items", [])),
                "by_category": cats,
            }

        # PC Parts
        parts_path = Path(data_sources.get("pc_parts", {}).get("path", "../pc_parts_agent/data"))
        parts_files = sorted(parts_path.glob("prices_*.json"))
        if parts_files:
            with open(parts_files[-1], encoding="utf-8") as f:
                parts_data = json.load(f)
            aggregated["pc_parts"]["items"] = parts_data.get("items", [])
            aggregated["pc_parts"]["summary"] = {
                "total": len(parts_data.get("items", [])),
            }

        # MSI Updates
        msi_path = Path(data_sources.get("msi_monitor", {}).get("path", "../msi_monitor_agent/data"))
        msi_files = sorted(msi_path.glob("msi_updates_*.json"))
        if msi_files:
            with open(msi_files[-1], encoding="utf-8") as f:
                msi_data = json.load(f)
            aggregated["msi_updates"]["items"] = msi_data.get("items", [])
            aggregated["msi_updates"]["summary"] = {
                "total": len(msi_data.get("items", [])),
            }

        self.data = aggregated
        logger.info(f"數據匯總完成: {aggregated}")
        return aggregated

    # ── Generation ─────────────────────────────────────────────

    def generate_html(self) -> str:
        """生成 HTML 主報告"""
        news = self.data.get("it_news", {})
        parts = self.data.get("pc_parts", {})
        msi = self.data.get("msi_updates", {})

        # AI summary
        summary_data = {
            "news_count": len(news.get("items", [])),
            "parts_count": len(parts.get("items", [])),
            "msi_count": len(msi.get("items", [])),
            "news_categories": news.get("summary", {}).get("by_category", {}),
        }
        ai_summary = self.analyzer.generate_insights(summary_data)

        # News section
        news_html = ""
        cats = news.get("summary", {}).get("by_category", {})
        for cat, items in self._group_news(news.get("items", [])).items():
            cat_items = ""
            for item in items[:5]:
                cat_items += f'<li><a href="{item["url"]}" target="_blank">{item["title"]}</a> <span class="source">({item["source"]})</span></li>'
            news_html += f'<div class="cat-block"><h3>{cat} ({len(items)})</h3><ul>{cat_items}</ul></div>'

        # Parts section
        parts_html = ""
        brands = {}
        for item in parts.get("items", []):
            b = item.get("brand", "Other")
            brands[b] = brands.get(b, 0) + 1
        for brand, count in sorted(brands.items(), key=lambda x: -x[1])[:10]:
            parts_html += f'<li>{brand}: {count} 款產品</li>'

        # MSI section
        msi_html = ""
        for item in msi.get("items", [])[:10]:
            msi_html += f'<li><a href="{item["url"]}" target="_blank">{item["title"]}</a> <span class="tag">{item.get("update_type", "news")}</span></li>'

        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>AI 商業情報每日報告 {self.today_str}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; line-height: 1.6; color: #333; background: #f0f2f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(135deg, #0f2027 0%, #203a43 50%, #2c5364 100%); color: white; padding: 40px 30px; border-radius: 12px; margin-bottom: 25px; }}
        .header h1 {{ font-size: 28px; margin-bottom: 10px; }}
        .ai-summary {{ background: #e3f2fd; border-left: 5px solid #1976d2; padding: 15px 20px; margin: 15px 0; border-radius: 4px; font-size: 14px; }}
        .kpi-row {{ display: flex; gap: 15px; margin-top: 20px; }}
        .kpi {{ background: rgba(255,255,255,0.15); padding: 15px 25px; border-radius: 8px; text-align: center; flex: 1; }}
        .kpi .num {{ font-size: 32px; font-weight: bold; }}
        .kpi .label {{ font-size: 13px; opacity: 0.8; }}
        .card {{ background: white; border-radius: 12px; padding: 25px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
        .card h2 {{ color: #0f2027; font-size: 20px; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 3px solid #2c5364; }}
        .card-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; }}
        .cat-block {{ background: #f8f9fa; padding: 15px; border-radius: 8px; margin-bottom: 10px; }}
        .cat-block h3 {{ color: #2c5364; margin-bottom: 10px; font-size: 15px; }}
        .cat-block ul {{ list-style: none; }}
        .cat-block li {{ padding: 6px 0; border-bottom: 1px solid #eee; font-size: 14px; }}
        .cat-block a {{ color: #1976d2; text-decoration: none; }}
        .cat-block a:hover {{ text-decoration: underline; }}
        .source {{ color: #888; font-size: 12px; }}
        .tag {{ background: #e8f5e9; color: #2e7d32; padding: 2px 8px; border-radius: 12px; font-size: 11px; }}
        ul {{ list-style: none; }}
        li {{ padding: 6px 0; border-bottom: 1px solid #eee; }}
        li:last-child {{ border-bottom: none; }}
        a {{ color: #1976d2; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        .footer {{ text-align: center; padding: 30px; color: #888; font-size: 13px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🤖 AI 商業情報每日報告</h1>
            <p>{self.today_str} | 整合 IT News / PC Parts / MSI 監控</p>
            <div class="ai-summary">
                <strong>🤖 AI 洞察:</strong> {ai_summary}
            </div>
            <div class="kpi-row">
                <div class="kpi"><div class="num">{summary_data['news_count']}</div><div class="label">IT 新聞</div></div>
                <div class="kpi"><div class="num">{summary_data['parts_count']}</div><div class="label">PC 零組件</div></div>
                <div class="kpi"><div class="num">{summary_data['msi_count']}</div><div class="label">MSI 更新</div></div>
            </div>
        </div>

        <div class="card-grid">
            <div class="card">
                <h2>📰 IT 行業新聞</h2>
                {news_html or '<p>暫無數據</p>'}
            </div>
            <div class="card">
                <h2>💻 PC 零組件</h2>
                <p>共 {parts.get("summary", {}).get("total", 0)} 款產品</p>
                <ul>{parts_html or '<li>暫無數據</li>'}</ul>
            </div>
            <div class="card">
                <h2>🎮 MSI 動態</h2>
                <ul>{msi_html or '<li>暫無數據</li>'}</ul>
            </div>
        </div>

        <div class="footer">
            <p>由 AI Business Intelligence Platform - Report Agent 自動生成</p>
            <p>生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>"""

        path = self.output_dir / f"master_report_{self.today_str}.html"
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        logger.info(f"HTML 報告已生成: {path}")
        return str(path)

    def generate_excel(self) -> str:
        """生成 Excel 工作簿"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # News sheet
            news = self.data.get("it_news", {})
            ws_news = wb.create_sheet("IT 新聞")
            ws_news.append(["標題", "來源", "分類", "URL", "摘要"])
            for item in news.get("items", []):
                ws_news.append([
                    item.get("title", ""),
                    item.get("source", ""),
                    item.get("category", ""),
                    item.get("url", ""),
                    item.get("summary", "")[:200],
                ])

            # Parts sheet
            parts = self.data.get("pc_parts", {})
            ws_parts = wb.create_sheet("PC 零組件")
            ws_parts.append(["品牌", "類別", "型號", "價格", "來源", "URL"])
            for item in parts.get("items", []):
                ws_parts.append([
                    item.get("brand", ""),
                    item.get("category", ""),
                    item.get("model_name", ""),
                    item.get("price", ""),
                    item.get("source", ""),
                    item.get("url", ""),
                ])

            # MSI sheet
            msi = self.data.get("msi_updates", {})
            ws_msi = wb.create_sheet("MSI 更新")
            ws_msi.append(["標題", "類型", "來源", "URL", "發布時間"])
            for item in msi.get("items", []):
                ws_msi.append([
                    item.get("title", ""),
                    item.get("update_type", ""),
                    item.get("source", ""),
                    item.get("url", ""),
                    item.get("published_at", ""),
                ])

            path = self.output_dir / f"master_report_{self.today_str}.xlsx"
            wb.save(str(path))
            logger.info(f"Excel 報告已生成: {path}")
            return str(path)
        except ImportError:
            logger.error("openpyxl 未安裝")
            return ""

    def generate_lark_doc(self) -> Optional[str]:
        """生成飛書雲文檔"""
        lark_cfg = self.cfg.get("distribution", {}).get("lark", {})
        if not lark_cfg.get("enabled"):
            return None

        client = LarkClient(lark_cfg)
        title = f"AI 商業情報報告 {self.today_str}"
        doc_token = client.create_doc(title)
        if not doc_token:
            return None

        # Build content blocks
        blocks = []

        # Title block
        blocks.append({
            "block_type": 2,
            "text": {"elements": [{"text_run": {"content": title}}], "style": {}}
        })

        # News summary
        news = self.data.get("it_news", {})
        blocks.append({
            "block_type": 2,
            "text": {"elements": [{"text_run": {"content": f"📰 IT 新聞 ({news.get('summary', {}).get('total', 0)} 條)"}}], "style": {}}
        })

        for item in news.get("items", [])[:10]:
            blocks.append({
                "block_type": 2,
                "text": {"elements": [{"text_run": {"content": f"• {item.get('title', '')} [{item.get('source', '')}]"}}], "style": {}}
            })

        blocks.append({
            "block_type": 2,
            "text": {"elements": [{"text_run": {"content": f"💻 PC 零組件 ({len(news.get('items', []))} 款)"}}], "style": {}}
        })

        for item in self.data.get("pc_parts", {}).get("items", [])[:10]:
            blocks.append({
                "block_type": 2,
                "text": {"elements": [{"text_run": {"content": f"• {item.get('brand', '')} {item.get('model_name', '')[:50]} ¥{item.get('price', '')}"}}], "style": {}}
            })

        client.batch_add_doc_blocks(doc_token, blocks)
        logger.info(f"飛書文檔已創建: {doc_token}")
        return doc_token

    def distribute(self) -> bool:
        """分發報告"""
        dist_cfg = self.cfg.get("distribution", {})
        email_cfg = dist_cfg.get("email", {})
        lark_cfg = dist_cfg.get("lark", {})

        html_path = self.output_dir / f"master_report_{self.today_str}.html"
        if not html_path.exists():
            logger.error("HTML 報告不存在")
            return False

        with open(html_path) as f:
            html = f.read()

        success = True

        # Email
        if email_cfg.get("enabled") and email_cfg.get("recipients"):
            sender = EmailSender(email_cfg)
            ok = sender.send(
                subject=f"AI 商業情報每日報告 {self.today_str}",
                html_content=html,
                recipients=email_cfg["recipients"],
            )
            success = success and ok

        # Lark
        if lark_cfg.get("enabled"):
            doc_token = self.generate_lark_doc()
            if doc_token:
                logger.info(f"飛書文檔已創建: {doc_token}")
            else:
                success = False

        return success

    def _group_news(self, items: List[Dict]) -> Dict[str, List[Dict]]:
        groups = {}
        for item in items:
            cat = item.get("category", "General")
            if cat not in groups:
                groups[cat] = []
            groups[cat].append(item)
        return groups

    def run(self, mode: str = "full") -> Dict:
        logger.info(f"Report Agent 啟動，模式: {mode}")
        result = {"status": "success"}

        try:
            if mode in ("full", "aggregate"):
                self.aggregate_all()

            if mode in ("full", "html"):
                self.aggregate_all()
                html_path = self.generate_html()
                result["html_path"] = html_path

            if mode in ("full", "excel"):
                if not self.data:
                    self.aggregate_all()
                excel_path = self.generate_excel()
                result["excel_path"] = excel_path

            if mode == "distribute":
                self.aggregate_all()
                self.generate_html()
                ok = self.distribute()
                result["distributed"] = ok

        except Exception as e:
            logger.error(f"執行失敗: {e}")
            result["status"] = "failed"
            result["error"] = str(e)

        return result
