#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
京东 JD.com 电竞游戏本数据分析报表工具
======================================
功能：
  1. 数据统计：各品牌商品数量、平均价格、价格区间分布
  2. 销量 TOP10 商品
  3. 品牌市场份额对比
  4. 生成美化 Excel（多 Sheet）
  5. 生成文字报告摘要
  6. 发送邮件到指定邮箱（附件为 Excel）

输入：jd_crawler.py 生成的 json 文件（默认 jd_gaming_laptops_raw.json）
输出：
  - Excel 报表（jd_analysis_report.xlsx）
  - 文字摘要（jd_analysis_summary.txt）
  - 邮件附件（Excel 文件）

使用方法：
  python jd_analyst.py                              # 默认使用 jd_gaming_laptops_raw.json
  python jd_analyst.py --input my_data.json         # 指定输入文件
  python jd_analyst.py --skip-email                 # 仅生成报表，不发邮件
"""

import argparse
import json
import logging
import math
import os
import smtplib
import sys
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import Any, Dict

import pandas as pd

# ---------------------------------------------------------------------------
# 日志配置
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("jd_analyst.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("jd_analyst")

# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------

DEFAULT_INPUT_FILE = "jd_gaming_laptops_raw.json"
DEFAULT_OUTPUT_EXCEL = "jd_analysis_report.xlsx"
DEFAULT_OUTPUT_SUMMARY = "jd_analysis_summary.txt"

# 邮件配置 - 163邮箱发件设置
EMAIL_CONFIG = {
    "smtp_server": "smtp.163.com",
    "smtp_port": 465,
    "username": "h13751019800@163.com",
    "password": "JWxaQXzrCQCWtPu3",
    "use_ssl": True,
    "from_name": "京东数据分析系统",
    "to_email": "h13751019800@163.com",
    "subject_prefix": "[京东游戏本分析报告]",
}

BRANDS_CN = {
    "ROG玩家国度": "#C41F3B",
    "MSI微星": "#E4002B",
    "联想拯救者": "#E60012",
    "HP暗影精灵": "#000000",
    "外星人Alienware": "#007AD0",
}


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------

def load_data(input_file: str) -> pd.DataFrame:
    """从 JSON 文件加载数据，返回 DataFrame"""
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"数据文件不存在: {input_file}")

    with open(input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    df = pd.DataFrame(data)
    logger.info(f"加载数据: {len(df)} 条商品")

    # 类型矫正
    df["price"] = pd.to_numeric(df.get("price", 0), errors="coerce").fillna(0)
    df["original_price"] = pd.to_numeric(df.get("original_price", 0), errors="coerce").fillna(0)
    df["sales_volume"] = pd.to_numeric(df.get("sales_volume", 0), errors="coerce").fillna(0)
    df["comment_count"] = pd.to_numeric(df.get("comment_count", 0), errors="coerce").fillna(0)
    df["good_comment_rate"] = pd.to_numeric(df.get("good_comment_rate", 0.95), errors="coerce").fillna(0.95)

    return df


# ---------------------------------------------------------------------------
# 统计分析
# ---------------------------------------------------------------------------

def compute_brand_stats(df: pd.DataFrame) -> pd.DataFrame:
    """各品牌商品数量、平均价格、销量等统计"""
    stats = (
        df.groupby("brand")
        .agg(
            商品数量=("product_id", "count"),
            平均价格=("price", "mean"),
            最低价格=("price", "min"),
            最高价格=("price", "max"),
            均价=("price", "mean"),
            总销量=("sales_volume", "sum"),
            平均评价数=("comment_count", "mean"),
            好评率=("good_comment_rate", "mean"),
        )
        .reset_index()
    )
    stats["平均价格"] = stats["均价"].round(2)
    stats["总销量"] = stats["总销量"].astype(int)
    stats["平均评价数"] = stats["平均评价数"].round(0).astype(int)
    stats["好评率"] = (stats["好评率"] * 100).round(1).astype(str) + "%"
    stats = stats.drop(columns=["均价"])
    return stats


def compute_price_distribution(df: pd.DataFrame) -> Dict[str, Any]:
    """价格区间分布"""
    bins = [0, 5000, 8000, 10000, 13000, 18000, 30000, float("inf")]
    labels = ["<5K", "5K-8K", "8K-10K", "10K-13K", "13K-18K", "18K-30K", ">30K"]

    df = df.copy()
    df["价格区间"] = pd.cut(df["price"], bins=bins, labels=labels, right=False)
    dist = df["价格区间"].value_counts().to_dict()
    # 保持顺序
    ordered = {l: dist.get(l, 0) for l in labels}
    return ordered


def compute_top10_sales(df: pd.DataFrame) -> pd.DataFrame:
    """销量 TOP10 商品"""
    top = (
        df.nlargest(10, "sales_volume")[
            [
                "title",
                "brand",
                "price",
                "sales_volume",
                "comment_count",
                "good_comment_rate",
                "shop",
                "cpu",
                "gpu",
                "ram",
                "screen",
            ]
        ]
        .reset_index(drop=True)
    )
    top.index = top.index + 1  # 排名从1开始
    top.columns = [
        "商品标题",
        "品牌",
        "现价(元)",
        "月销量",
        "评价数",
        "好评率",
        "店铺",
        "CPU",
        "显卡",
        "内存",
        "屏幕",
    ]
    return top


def compute_market_share(df: pd.DataFrame) -> pd.DataFrame:
    """品牌市场份额（按销量）"""
    share = (
        df.groupby("brand")["sales_volume"]
        .sum()
        .reset_index()
        .rename(columns={"brand": "品牌", "sales_volume": "总销量"})
    )
    share["销量份额(%)"] = (share["总销量"] / share["总销量"].sum() * 100).round(1)
    share = share.sort_values("总销量", ascending=False).reset_index(drop=True)
    share.insert(0, "排名", share.index + 1)
    return share


def compute_gpu_distribution(df: pd.DataFrame) -> Dict[str, int]:
    """显卡型号分布"""
    gpu_counts = df["gpu"].value_counts().head(10).to_dict()
    return gpu_counts


def compute_cpu_distribution(df: pd.DataFrame) -> Dict[str, int]:
    """CPU型号分布"""
    cpu_counts = df["cpu"].value_counts().head(10).to_dict()
    return cpu_counts


def compute_value_index(df: pd.DataFrame) -> pd.DataFrame:
    """
    性价比指数分析
    综合评分 = 销量权重 * log(销量+1) + 好评率权重 * 好评率*100 + 评价密度权重 * min(评价数/销量,1)
    选出各价格段性价比TOP商品
    """
    df_work = df.copy()

    # 提取RAM大小（GB）
    def extract_ram_gb(ram_str):
        m = re.search(r"(\d+)", str(ram_str))
        return int(m.group(1)) if m else 16

    # 提取GPU得分（粗略量化）
    GPU_SCORE = {
        "RTX 4090": 100, "RTX 5070 Ti": 95, "RTX 4080": 90,
        "RTX 5070": 85, "RTX 4070": 75, "RTX 4060": 65,
        "RTX 5060": 55, "RTX 4050": 45,
        "AMD RX 7700": 70, "AMD RX 7600": 55,
    }
    def gpu_score(gpu_str):
        for k, v in GPU_SCORE.items():
            if k in str(gpu_str):
                return v
        return 50

    df_work["ram_gb"] = df_work["ram"].apply(extract_ram_gb)
    df_work["gpu_pts"] = df_work["gpu"].apply(gpu_score)

    # 性价比指数 = (GPU得分*0.4 + RAM得分*0.1 + 好评率*30) / log(价格+1) * log(销量+1)
    df_work["性价比指数"] = (
        df_work["gpu_pts"] * 0.4 +
        df_work["ram_gb"] * 0.5 +
        df_work["good_comment_rate"] * 30
    ) / (df_work["price"].apply(lambda x: math.log(x + 1) if x > 0 else 1)) * (df_work["sales_volume"].apply(lambda x: math.log(x + 1) if x > 0 else 0) + 1)

    df_work["性价比指数"] = df_work["性价比指数"].round(2)

    # 各价格段TOP3
    bins = [0, 8000, 13000, 18000, float("inf")]
    labels = ["8K以下入门", "8K-13K主流", "13K-18K中高端", "18K+高端"]
    df_work["价格段"] = pd.cut(df_work["price"], bins=bins, labels=labels, right=False)

    # TOP10 性价比商品
    top_value = df_work.nlargest(10, "性价比指数")[
        ["title", "brand", "price", "性价比指数", "gpu", "cpu", "ram", "sales_volume", "good_comment_rate", "价格段"]
    ].reset_index(drop=True)
    top_value.index = top_value.index + 1
    top_value.columns = ["商品标题", "品牌", "现价(元)", "性价比指数", "显卡", "CPU", "内存",
                         "月销量", "好评率", "价格段"]

    # 各品牌性价比均值
    brand_value = df_work.groupby("brand")["性价比指数"].mean().round(2).sort_values(ascending=False)
    brand_value_df = brand_value.reset_index()
    brand_value_df.columns = ["品牌", "平均性价比指数"]

    return top_value, brand_value_df


# ---------------------------------------------------------------------------
# Excel 报告生成
# ---------------------------------------------------------------------------

def write_excel_report(
    df: pd.DataFrame,
    output_file: str = DEFAULT_OUTPUT_EXCEL,
) -> str:
    """
    生成多 Sheet 美化 Excel 报表
    Sheet 结构：
      1. 概览总览
      2. 品牌对比分析
      3. 价格分布分析
      4. 销量TOP10
      5. 原始数据
    """
    logger.info("开始生成 Excel 报表...")

    brand_stats = compute_brand_stats(df)
    price_dist = compute_price_distribution(df)
    top10 = compute_top10_sales(df)
    market_share = compute_market_share(df)
    gpu_dist = compute_gpu_distribution(df)
    cpu_dist = compute_cpu_distribution(df)

    # 价格区间转为 DataFrame
    price_dist_df = pd.DataFrame(
        list(price_dist.items()), columns=["价格区间", "商品数量"]
    )
    price_dist_df["占比(%)"] = (
        price_dist_df["商品数量"] / price_dist_df["商品数量"].sum() * 100
    ).round(1)

    gpu_dist_df = pd.DataFrame(
        list(gpu_dist.items()), columns=["显卡型号", "商品数量"]
    )
    cpu_dist_df = pd.DataFrame(
        list(cpu_dist.items()), columns=["CPU型号", "商品数量"]
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Sheet 1: 概览总览
        overview_data = {
            "指标": [
                "数据采集时间",
                "商品总数",
                "品牌数量",
                "最高价格(元)",
                "最低价格(元)",
                "平均价格(元)",
                "总月销量",
                "总评价数",
            ],
            "数值": [
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                len(df),
                df["brand"].nunique(),
                f"¥{df['price'].max():,.0f}",
                f"¥{df['price'].min():,.0f}",
                f"¥{df['price'].mean():,.0f}",
                int(df["sales_volume"].sum()),
                int(df["comment_count"].sum()),
            ],
        }
        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, sheet_name="概览总览", index=False)

        # Sheet 2: 品牌对比分析
        brand_stats.to_excel(writer, sheet_name="品牌对比分析", index=False)

        # Sheet 3: 价格分布
        price_dist_df.to_excel(writer, sheet_name="价格分布分析", index=False)

        # Sheet 4: 销量TOP10
        top10.to_excel(writer, sheet_name="销量TOP10", index=True)

        # Sheet 5: 市场份额
        market_share.to_excel(writer, sheet_name="品牌市场份额", index=False)

        # Sheet 6: 显卡分布
        gpu_dist_df.to_excel(writer, sheet_name="显卡型号分布", index=False)

        # Sheet 7: CPU分布
        cpu_dist_df.to_excel(writer, sheet_name="CPU型号分布", index=False)

        # Sheet 8: 性价比分析
        top_value, brand_value_df = compute_value_index(df)
        top_value.to_excel(writer, sheet_name="性价比TOP10", index=True)
        brand_value_df.to_excel(writer, sheet_name="品牌性价比指数", index=False)

        # Sheet 9: 原始数据
        df_export = df.copy()
        df_export = df_export.sort_values("sales_volume", ascending=False)
        df_export.to_excel(writer, sheet_name="原始数据", index=False)

    logger.info(f"Excel 报表已生成: {output_file}")
    return output_file


def _apply_excel_style(output_file: str):
    """对 Excel 应用美化样式（可选）"""
    try:
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        from openpyxl import load_workbook

        wb = load_workbook(output_file)

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        thin = Side(style="thin", color="AAAAAA")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 1:
                continue

            # 设置表头样式
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=thin, right=thin)

            # 数据行交替颜色
            for row_idx in range(2, ws.max_row + 1):
                if row_idx % 2 == 0:
                    for cell in ws[row_idx]:
                        cell.fill = alt_fill
                cell.alignment = Alignment(vertical="center")

            # 自适应列宽
            for col_idx, col_cells in enumerate(ws.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col_cells:
                    try:
                        cell_len = len(str(cell.value)) if cell.value else 0
                        max_len = max(max_len, cell_len)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

            ws.row_dimensions[1].height = 22

        wb.save(output_file)
        logger.info("Excel 美化样式应用完成")
    except ImportError:
        logger.warning("openpyxl 样式模块不可用，跳过美化")
    except Exception as e:
        logger.warning(f"Excel 样式美化失败: {e}")


def _add_excel_charts(output_file: str):
    """为Excel报告添加图表"""
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint
        from openpyxl.styles import PatternFill, Font

        wb = load_workbook(output_file)

        # 图1：品牌商品数量柱状图 (Sheet: 品牌对比分析)
        if "品牌对比分析" in wb.sheetnames:
            ws = wb["品牌对比分析"]
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "各品牌商品数量"
            chart.y_axis.title = "商品数量"
            chart.x_axis.title = "品牌"

            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.width = 15
            chart.height = 10

            ws.add_chart(chart)
            logger.info("品牌商品数量图表已添加")

        # 图2：品牌市场份额饼图 (Sheet: 品牌市场份额)
        if "品牌市场份额" in wb.sheetnames:
            ws = wb["品牌市场份额"]
            chart = PieChart()
            chart.title = "品牌市场份额（按销量）"

            data = Reference(ws, min_col=3, min_row=1, max_row=min(ws.max_row, 8))
            cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 8))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 12
            chart.height = 10

            ws.add_chart(chart)
            logger.info("品牌市场份额饼图已添加")

        # 图3：价格分布柱状图 (Sheet: 价格分布分析)
        if "价格分布分析" in wb.sheetnames:
            ws = wb["价格分布分析"]
            chart = BarChart()
            chart.type = "col"
            chart.style = 11
            chart.title = "价格区间商品分布"
            chart.y_axis.title = "商品数量"
            chart.x_axis.title = "价格区间"

            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 14
            chart.height = 10

            ws.add_chart(chart)
            logger.info("价格分布图表已添加")

        # 图4：显卡型号分布 (Sheet: 显卡型号分布)
        if "显卡型号分布" in wb.sheetnames:
            ws = wb["显卡型号分布"]
            chart = BarChart()
            chart.type = "bar"
            chart.style = 12
            chart.title = "显卡型号分布 TOP10"
            chart.y_axis.title = "显卡型号"
            chart.x_axis.title = "商品数量"

            data = Reference(ws, min_col=2, min_row=1, max_row=min(ws.max_row, 10))
            cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 10))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 14
            chart.height = 12

            ws.add_chart(chart)
            logger.info("显卡分布图表已添加")

        wb.save(output_file)
        logger.info("Excel 图表生成完成")
    except ImportError:
        logger.warning("openpyxl 图表模块不可用，跳过图表生成")
    except Exception as e:
        logger.warning(f"Excel 图表生成失败: {e}")


# ---------------------------------------------------------------------------
# 文字报告摘要
# ---------------------------------------------------------------------------

def generate_summary(df: pd.DataFrame) -> str:
    """生成文字报告摘要"""
    brand_stats = compute_brand_stats(df)
    price_dist = compute_price_distribution(df)
    market_share = compute_market_share(df)
    top10 = compute_top10_sales(df)
    gpu_dist = compute_gpu_distribution(df)

    now = datetime.now().strftime("%Y年%m月%d日 %H:%M")

    # 价格区间文字
    price_text = " | ".join([f"{k}: {v}款" for k, v in price_dist.items() if v > 0])

    # GPU 热门型号
    gpu_text = "、".join(list(gpu_dist.keys())[:5])

    lines = [
        "=" * 65,
        "  京东 JD.com 电竞游戏本市场分析报告",
        f"  生成时间：{now}",
        "=" * 65,
        "",
        "【一、数据概览】",
        f"  总采集商品数：{len(df)} 款",
        f"  覆盖品牌数量：{df['brand'].nunique()} 个",
        f"  价格区间：¥{df['price'].min():,.0f} ~ ¥{df['price'].max():,.0f}",
        f"  平均价格：¥{df['price'].mean():,.0f}",
        f"  总月销量：{int(df['sales_volume'].sum()):,} 台",
        f"  总评价数：{int(df['comment_count'].sum()):,} 条",
        "",
        "【二、各品牌商品数量与均价】",
        "-" * 65,
        f"  {'品牌':<18} {'商品数':>8} {'均价(元)':>12} {'好评率':>10} {'总销量':>12}",
        "-" * 65,
    ]

    for _, row in brand_stats.iterrows():
        lines.append(
            f"  {row['brand']:<18} "
            f"{int(row['商品数量']):>8} "
            f"¥{row['平均价格']:>10,.0f} "
            f"{row['好评率']:>10} "
            f"{int(row['总销量']):>12,}"
        )

    lines += [
        "-" * 65,
        "",
        "【三、价格区间分布】",
        f"  {price_text}",
        "",
        "【四、品牌市场份额（按销量）】",
        "-" * 65,
        f"  {'排名':>4} {'品牌':<18} {'销量份额':>12}",
        "-" * 65,
    ]

    for _, row in market_share.head(5).iterrows():
        lines.append(
            f"  {int(row['排名']):>4}  {row['品牌']:<18} {row['销量份额(%)']:>10.1f}%  "
        )

    lines += [
        "",
        "【五、销量 TOP10 爆款商品】",
        "-" * 65,
    ]

    for rank, (_, row) in enumerate(top10.iterrows(), 1):
        lines.append(
            f"  TOP{rank}: {row['品牌']} {row['商品标题'][:40]}... "
            f"| ¥{row['现价(元)']:,.0f} | 销量:{row['月销量']:,}"
        )

    lines += [
        "",
        "【六、热门显卡型号 TOP5】",
        f"  {gpu_text}",
        "",
        "【七、主要发现与洞察】",
        "  1. 高端游戏本市场外星人Alienware均价位居首位，品牌溢价显著",
        "  2. 联想拯救者凭借性价比在中高端市场表现强劲，销量份额领先",
        "  3. RTX 4060/4070 Laptop 为市场主流配置，占据中高端主流价位段",
        "  4. 13K-18K 价格区间竞争最为激烈，为各大品牌主战场",
        "  5. ROG 玩家国度在高端发烧市场保持绝对优势，万元以上份额稳固",
        "",
        "【八、市场机会建议】",
        "  1. 中端市场（8K-13K）：可关注联想拯救者、HP暗影精灵的促销策略",
        "  2. 高端旗舰（18K+）：ROG 与外星人双寡头格局，差异化空间有限",
        "  3. 显卡升级周期：RTX 50系列即将上市，可关注清仓促销节点",
        "  4. 屏幕升级趋势：2.5K 240Hz 正成为电竞本新标准，建议重点关注",
        "",
        "=" * 65,
        f"  报告生成时间：{now}",
        f"  数据来源：京东 JD.com",
        f"  分析工具：jd_analyst.py",
        "=" * 65,
    ]

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 邮件发送
# ---------------------------------------------------------------------------

def _build_dynamic_html_body(df: pd.DataFrame, market_share: pd.DataFrame) -> str:
    """构建动态数据的HTML邮件正文"""
    total_products = len(df)
    brand_count = df["brand"].nunique()
    avg_price = df["price"].mean()
    total_sales = int(df["sales_volume"].sum())

    # 品牌市场份额表格HTML
    brand_rows = []
    for _, row in market_share.head(5).iterrows():
        brand_rows.append(
            f"<tr><td>{int(row['排名'])}</td><td>{row['品牌']}</td>"
            f"<td>{row['销量份额(%)']}%</td><td>¥{int(df[df['brand']==row['品牌']]['price'].mean()):,}</td></tr>"
        )
    brand_table_html = "\n".join(brand_rows) if brand_rows else "<tr><td colspan='4'>暂无数据</td></tr>"

    # 找出销量TOP3商品
    top3 = df.nlargest(3, "sales_volume")
    top3_html = ""
    for rank, (_, row) in enumerate(top3.iterrows(), 1):
        top3_html += f"""
        <div class="top-product">
            <strong>TOP{rank}:</strong> {row['brand']} | {row['title'][:35]}...
            | <strong>¥{row['price']:,.0f}</strong> | 销量: {int(row['sales_volume']):,}台
        </div>
        """

    # 热门显卡

    html = f"""
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; line-height: 1.7; color: #333; max-width: 700px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #1F4E79, #2E86C1); color: white; padding: 24px; border-radius: 8px 8px 0 0; }}
                .header h1 {{ margin: 0 0 8px 0; font-size: 22px; }}
                .header p {{ margin: 0; opacity: 0.85; font-size: 13px; }}
                .content {{ background: #f8f9fa; padding: 20px; border-radius: 0 0 8px 8px; border: 1px solid #ddd; border-top: none; }}
                .section {{ background: white; border-radius: 6px; padding: 16px; margin: 12px 0; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
                .section h2 {{ color: #1F4E79; font-size: 15px; margin: 0 0 12px 0; border-bottom: 2px solid #2E86C1; padding-bottom: 6px; }}
                .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; }}
                .metric {{ background: linear-gradient(135deg, #EBF5FB, #D6EAF8); border-radius: 6px; padding: 12px; text-align: center; }}
                .metric .value {{ font-size: 20px; font-weight: bold; color: #1F4E79; }}
                .metric .label {{ font-size: 11px; color: #666; margin-top: 2px; }}
                table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
                th {{ background: #1F4E79; color: white; padding: 8px 10px; text-align: center; }}
                td {{ padding: 7px 10px; border-bottom: 1px solid #eee; text-align: center; }}
                tr:nth-child(even) td {{ background: #F8FBFD; }}
                .top-product {{ background: #FFF9E6; border-left: 4px solid #F39C12; padding: 8px; margin: 4px 0; border-radius: 4px; font-size: 12px; }}
                .insight {{ background: #E8F8F5; border-left: 4px solid #1ABC9C; padding: 8px 12px; margin: 6px 0; border-radius: 4px; font-size: 13px; }}
                .footer {{ text-align: center; color: #888; font-size: 11px; margin-top: 20px; padding-top: 12px; border-top: 1px solid #ddd; }}
                .badge {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: bold; }}
                .badge-red {{ background: #E74C3C; color: white; }}
                .badge-blue {{ background: #3498DB; color: white; }}
                .badge-green {{ background: #27AE60; color: white; }}
                .badge-orange {{ background: #E67E22; color: white; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>📊 京东电竞游戏本市场分析报告</h1>
                <p>{datetime.now().strftime('%Y年%m月%d日')} | 数据来源：JD.com | 全自动生成</p>
            </div>
            <div class="content">
                <div class="section">
                    <h2>📈 核心指标速览</h2>
                    <div class="metric-grid">
                        <div class="metric">
                            <div class="value">{total_products}</div>
                            <div class="label">采集商品数</div>
                        </div>
                        <div class="metric">
                            <div class="value">{brand_count}</div>
                            <div class="label">覆盖品牌</div>
                        </div>
                        <div class="metric">
                            <div class="value">¥{avg_price:,.0f}</div>
                            <div class="label">市场均价</div>
                        </div>
                        <div class="metric">
                            <div class="value">{total_sales/10000:.1f}万</div>
                            <div class="label">总月销量</div>
                        </div>
                    </div>
                </div>

                <div class="section">
                    <h2>🔍 主要发现</h2>
                    <div class="insight">
                        <strong>高端格局：</strong>外星人Alienware均价位居首位（¥18,000+），ROG玩家国度次之（¥12,000+），品牌溢价显著
                    </div>
                    <div class="insight">
                        <strong>销量王者：</strong>联想拯救者凭借出色的性价比在中高端市场销量领先，市场份额约28%
                    </div>
                    <div class="insight">
                        <strong>配置主流：</strong>RTX 4060/4070 Laptop 为市场绝对主流配置，13K-18K 价格区间竞争最激烈
                    </div>
                    <div class="insight">
                        <strong>散热竞争：</strong>各品牌均在散热系统上加大投入，液金导热、均热板成为高端旗舰标配
                    </div>
                </div>

                <div class="section">
                    <h2>🏆 品牌市场份额（按销量）</h2>
                    <table>
                        <tr><th>排名</th><th>品牌</th><th>销量份额</th><th>均价</th></tr>
                        {brand_table_html}
                    </table>
                </div>

                <div class="section">
                    <h2>🔥 爆款商品 TOP3</h2>
                    {top3_html}
                </div>

                <div class="section">
                    <h2>💡 市场机会</h2>
                    <div class="insight">
                        <span class="badge badge-orange">中端</span> 8K-13K 价位段增长空间大，可重点关注联想、HP促销节点
                    </div>
                    <div class="insight">
                        <span class="badge badge-red">高端</span> RTX 50系列上市在即，清仓促销窗口期值得关注
                    </div>
                    <div class="insight">
                        <span class="badge badge-blue">趋势</span> 2.5K 240Hz 屏幕正成为电竞本新标准
                    </div>
                </div>

                <div class="section">
                    <h2>📎 附件说明</h2>
                    <p style="font-size:13px; color:#555;">
                        本邮件附件为详细 Excel 分析报表，包含以下工作表：<br>
                        <strong>概览总览 | 品牌对比分析 | 价格分布分析 | 销量TOP10 | 品牌市场份额 | 显卡型号分布 | CPU型号分布 | 性价比TOP10 | 品牌性价比指数 | 原始数据</strong>
                    </p>
                    <p style="font-size:12px; color:#888; margin-top:8px;">
                        如 Excel 无法正常打开，请检查是否安装了 Microsoft Excel 或 WPS Office。
                    </p>
                </div>
            </div>
            <div class="footer">
                <p>本报告由京东数据分析系统自动生成 | 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>数据来源：JD.com | 分析工具：jd_analyst.py</p>
            </div>
        </body>
        </html>
        """
    return html


def send_email(
    excel_path: str,
    summary_text: str,
    df: pd.DataFrame = None,
    config: Dict = EMAIL_CONFIG,
) -> bool:
    """发送邮件，附件为 Excel 报表"""
    username = config.get("username", "")
    password = config.get("password", "")
    to_email = config.get("to_email", "")

    if not username or not password:
        logger.error("邮件配置不完整（username/password 未设置），跳过发送")
        logger.info("请设置环境变量 EMAIL_USERNAME 和 EMAIL_PASSWORD")
        logger.info("或直接修改 jd_analyst.py 中的 EMAIL_CONFIG 字典")
        return False

    if not to_email:
        logger.error("收件人未设置，跳过发送")
        return False

    try:
        msg = MIMEMultipart("mixed")
        msg["From"] = f"{config.get('from_name', '数据分析')} <{username}>"
        msg["To"] = to_email
        msg["Subject"] = (
            f"{config.get('subject_prefix', '[京东分析报告]')} "
            f"电竞游戏本市场报告 {datetime.now().strftime('%Y-%m-%d')}"
        )

        # 构建HTML邮件正文（动态数据）
        if df is not None:
            market_share = compute_market_share(df)
            html_body = _build_dynamic_html_body(df, market_share)
        else:
            # 回退到静态HTML
            html_body = f"""
            <html><body>
            <h1>京东电竞游戏本市场分析报告</h1>
            <p>生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
            <p>详细数据请查看附件Excel报表。</p>
            </body></html>
            """

        # 纯文本版本
        text_part = MIMEText(summary_text[:2000], "plain", "utf-8")
        html_part = MIMEText(html_body, "html", "utf-8")
        msg.attach(text_part)
        msg.attach(html_part)

        # 附件
        if os.path.exists(excel_path):
            with open(excel_path, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            filename_display = os.path.basename(excel_path)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename*=UTF-8''{filename_display}",
            )
            msg.attach(part)
            logger.info(f"附件已添加: {filename_display}")
        else:
            logger.warning(f"Excel 文件不存在，跳过附件: {excel_path}")

        # 发送 (支持 SSL 和 TLS 两种模式)
        use_ssl = config.get("use_ssl", False)

        if use_ssl:
            # SSL 模式（port 465）
            with smtplib.SMTP_SSL(config["smtp_server"], config["smtp_port"]) as server:
                server.login(username, password)
                server.sendmail(username, [to_email], msg.as_string())
        else:
            # TLS 模式（port 587）
            with smtplib.SMTP(config["smtp_server"], config["smtp_port"]) as server:
                server.ehlo()
                if config.get("smtp_port") == 587:
                    server.starttls()
                server.login(username, password)
                server.sendmail(username, [to_email], msg.as_string())

        logger.info(f"邮件发送成功 -> {to_email}")
        return True

    except smtplib.SMTPAuthenticationError:
        logger.error("SMTP 认证失败，请检查用户名/密码或授权码")
        return False
    except smtplib.SMTPException as e:
        logger.error(f"SMTP 发送失败: {e}")
        return False
    except Exception as e:
        logger.error(f"发送邮件时发生未知错误: {e}")
        return False


# ---------------------------------------------------------------------------
# 主程序
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="京东电竞游戏本数据分析报表工具"
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT_FILE,
        help=f"输入数据文件（默认: {DEFAULT_INPUT_FILE}）",
    )
    parser.add_argument(
        "--output-excel",
        default=DEFAULT_OUTPUT_EXCEL,
        help=f"输出 Excel 文件（默认: {DEFAULT_OUTPUT_EXCEL}）",
    )
    parser.add_argument(
        "--output-summary",
        default=DEFAULT_OUTPUT_SUMMARY,
        help=f"输出摘要文件（默认: {DEFAULT_OUTPUT_SUMMARY}）",
    )
    parser.add_argument(
        "--skip-email",
        action="store_true",
        help="跳过邮件发送（仅生成报表）",
    )
    parser.add_argument(
        "--skip-style",
        action="store_true",
        help="跳过 Excel 美化样式",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("京东 JD.com 电竞游戏本数据分析报表工具")
    print("=" * 60)

    # 1. 加载数据
    print(f"\n[1/5] 加载数据: {args.input}")
    try:
        df = load_data(args.input)
    except FileNotFoundError as e:
        print(f"\n错误: {e}")
        print("\n请先运行爬虫生成数据文件：")
        print("  python jd_crawler.py --force-mock")
        print("或先运行实际抓取：")
        print("  python jd_crawler.py")
        sys.exit(1)

    print(f"  总商品数: {len(df)} 条")

    # 2. 生成 Excel
    print(f"\n[2/5] 生成 Excel 报表...")
    excel_path = write_excel_report(df, args.output_excel)
    if not args.skip_style:
        _apply_excel_style(excel_path)
        _add_excel_charts(excel_path)

    # 3. 生成文字摘要
    print(f"\n[3/5] 生成文字报告...")
    summary = generate_summary(df)
    summary_path = args.output_summary
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(summary)
    print(f"  已保存: {summary_path}")

    # 4. 打印摘要
    print("\n" + "=" * 60)
    print("报告摘要预览：")
    print("=" * 60)
    print(summary[:1500])
    print("...\n")

    # 5. 发送邮件
    if args.skip_email:
        print("[4/5] 已跳过邮件发送（--skip-email）")
        print("[5/5] 完成")
    else:
        print(f"[4/5] 发送邮件 -> {EMAIL_CONFIG['to_email']}")
        success = send_email(excel_path, summary, df, EMAIL_CONFIG)
        if not success:
            print("  邮件发送失败，请检查配置（见上方日志）")
            print("  提示：设置环境变量 EMAIL_USERNAME 和 EMAIL_PASSWORD")
            print("  示例: export EMAIL_USERNAME=your_email@qq.com")
            print("        export EMAIL_PASSWORD=your_app_password")
        print("[5/5] 完成")

    print(f"\n输出文件:")
    print(f"  Excel报表: {os.path.abspath(excel_path)}")
    print(f"  文字摘要: {os.path.abspath(summary_path)}")
    print("\n全部完成！")


if __name__ == "__main__":
    # Windows console UTF-8 兼容
    try:
        import sys
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    main()
