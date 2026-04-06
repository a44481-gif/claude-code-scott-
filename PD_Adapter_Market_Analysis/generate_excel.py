#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PD 大電流適配器集成技術應用市場分析 - Excel數據文件生成腳本
Global PD High-Current Adapter Integration Technology Market Analysis - Excel Data Generator
生成日期: 2025年4月
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
import os

# ==================== 樣式定義 ====================
# 深色主題樣式
DARK_BLUE = 'FF1f3864'
DARK_NAVY = 'FF0d1117'
LIGHT_GRAY = 'FFf2f2f2'
WHITE = 'FFffffff'
ACCENT_BLUE = 'FF58a6ff'
ACCENT_GREEN = 'FF3fb950'
ACCENT_ORANGE = 'FFf0883e'
ACCENT_CYAN = 'FF00d4ff'

def get_header_fill():
    """深色表头填充"""
    return PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')

def get_alt_row_fill():
    """交替行填充"""
    return PatternFill(start_color='FFe8e8e8', end_color='FFe8e8e8', fill_type='solid')

def get_border():
    """细边框样式"""
    thin = Side(style='thin', color='FFcccccc')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def get_header_font():
    """表头字体"""
    return Font(name='Microsoft JhengHei', size=11, bold=True, color=WHITE)

def get_data_font():
    """数据字体"""
    return Font(name='Microsoft JhengHei', size=10)

def style_header_row(ws, row_num, col_count):
    """设置表头行样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = get_header_fill()
        cell.font = get_header_font()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = get_border()

def style_data_row(ws, row_num, col_count, is_alt=False):
    """设置数据行样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = get_data_font()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = get_border()
        if is_alt:
            cell.fill = get_alt_row_fill()

def auto_fit_columns(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

def freeze_top_row(ws):
    """冻结首行"""
    ws.freeze_panes = 'A2'

# ==================== 工作表1: 市場概覽 ====================
def create_market_overview_sheet(wb):
    """创建市场概览表"""
    ws = wb.create_sheet(title='市場概覽')

    # 表头
    headers = ['年份', '全球出貨量(百萬台)', '全球市場規模(億美元)', '同比增長(%)', '集成技術滲透率(%)']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    # 数据
    data = [
        [2025, 485, 14.2, None, 45],
        [2026, 572, 17.1, 20.4, 58],
        [2027, 672, 20.6, 20.5, 72],
        [2028, 778, 24.0, 16.5, 82],
        [2029, 878, 26.8, 11.7, 88],
        [2030, 978, 28.7, 7.1, 92],
    ]

    for i, row_data in enumerate(data):
        # 计算同比增长率
        if i > 0:
            prev_val = data[i-1][2]
            curr_val = row_data[2]
            yoy = round((curr_val - prev_val) / prev_val * 100, 1)
            row_data[3] = yoy
        ws.append(row_data)
        style_data_row(ws, i + 2, len(headers), is_alt=(i % 2 == 1))

    # 格式化
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 22

    # 数字格式
    for row in range(2, 8):
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).number_format = '0.0'
        ws.cell(row=row, column=4).number_format = '0.0'
        ws.cell(row=row, column=5).number_format = '0'

    freeze_top_row(ws)
    return ws

# ==================== 工作表2: 功率結構 ====================
def create_power_structure_sheet(wb):
    """创建功率结构表"""
    ws = wb.create_sheet(title='功率結構')

    # 2025功率结构
    ws['A1'] = '2025功率結構'
    ws['A1'].font = Font(name='Microsoft JhengHei', size=14, bold=True, color=DARK_BLUE)
    ws.merge_cells('A1:D1')

    headers_2025 = ['功率等級', '市場份額(%)', '同比增長(%)', 'CAGR(%)']
    ws.append([''] * 4)  # 空行
    ws.append(headers_2025)
    style_header_row(ws, 3, 4)

    power_2025_data = [
        ['65W', 52, 5.2, 8.2],
        ['100W', 28, 12.5, 16.5],
        ['140W', 12, 28.7, 28.7],
        ['240W', 8, 38.2, 32.4],
    ]
    for i, row_data in enumerate(power_2025_data):
        ws.append(row_data)
        style_data_row(ws, i + 4, 4, is_alt=(i % 2 == 1))

    # 2030功率结构
    ws['F1'] = '2030功率結構'
    ws['F1'].font = Font(name='Microsoft JhengHei', size=14, bold=True, color=DARK_BLUE)
    ws.merge_cells('F1:I1')

    headers_2030 = ['功率等級', '市場份額(%)', '同比增長(%)', 'CAGR(%)']
    ws.append([''] * 5)  # 空行
    ws.append([''] * 5)
    ws.append(headers_2030)
    style_header_row(ws, 3, 9)

    power_2030_data = [
        ['65W', 35, 3.8, 8.2],
        ['100W', 30, 15.2, 16.5],
        ['140W', 20, 30.5, 28.7],
        ['240W', 15, 42.1, 32.4],
    ]
    for i, row_data in enumerate(power_2030_data):
        for j, val in enumerate(row_data):
            ws.cell(row=i + 4, column=6 + j, value=val)
            style_data_row(ws, i + 4, 9, is_alt=(i % 2 == 1))

    # 格式化
    for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 16

    freeze_top_row(ws)
    return ws

# ==================== 工作表3: 區域市場 ====================
def create_regional_market_sheet(wb):
    """创建区域市场表"""
    ws = wb.create_sheet(title='區域市場')

    headers = ['地區', '2025市場份額(%)', '2030市場份額(%)', '2025市場規模(億美元)', 'CAGR(%)']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    regional_data = [
        ['中國', 48, 45, 6.82, 13.8],
        ['北美', 20, 19, 2.84, 14.2],
        ['歐洲', 17, 20, 2.41, 18.5],
        ['東南亞', 8, 10, 1.14, 22.0],
        ['日本韓國', 5, 4, 0.71, 12.5],
        ['其他', 2, 2, 0.28, 15.0],
    ]

    for i, row_data in enumerate(regional_data):
        ws.append(row_data)
        style_data_row(ws, i + 2, len(headers), is_alt=(i % 2 == 1))

    # 格式化
    for col_num, width in enumerate([12, 18, 18, 22, 12], 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row in range(2, 8):
        for col in [2, 3, 4, 5]:
            ws.cell(row=row, column=col).number_format = '0.0'
        ws.cell(row=row, column=2).number_format = '0'
        ws.cell(row=row, column=3).number_format = '0'
        ws.cell(row=row, column=5).number_format = '0.0'

    freeze_top_row(ws)
    return ws

# ==================== 工作表4: 技術方案對比 ====================
def create_tech_comparison_sheet(wb):
    """创建技术方案对比表"""
    ws = wb.create_sheet(title='技術方案對比')

    headers = ['方案名稱', '代表芯片', '集成度', '最高功率(W)', '效率(%)', '功率密度(W/in³)',
               'BOM元件數', 'PCB面積(65W)', '成本(美元)', '開發週期(月)', '主要優勢', '主要劣勢', '典型應用']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    tech_data = [
        ['PFC+LLC GaN合封', 'PI InnoSwitch4-CZ', '★★★★☆', 220, '>95', '25-30', '25-35', '8×8cm', '$4.5-6.0', '4-6', '最高可靠性', '複雜設計', '高端筆電'],
        ['ACF單芯片', 'DK8607/FAN6966', '★★★★☆', 120, '93-94.5', '18-22', '30-40', '10×10cm', '$3.0-4.5', '3-5', '性價比最優', '功率上限', '手機/strip充'],
        ['協議+DC-DC SoC', 'SC9712/SW3536', '★★★★★', 140, '92-94', '15-20', '15-20', '6×6cm', '$2.5-3.5', '2-4', '最小PCB面積', '熱管理挑戰', '多口桌面充'],
        ['全數字功率', 'UCC29950/JW7726', '★★★☆☆', 240, '>96', '28-35', '20-30', '9×9cm', '$5.0-8.0', '5-8', '智能控制最強', '成本最高', '工業/醫療'],
        ['離散方案(對比基準)', '分立元器件', '★☆☆☆☆', 240, '88-92', '10-15', '45-55', '15×15cm', '$6.5-8.5', '2-3', '設計簡單', 'BOM最多', '低端適配器'],
    ]

    for i, row_data in enumerate(tech_data):
        ws.append(row_data)
        style_data_row(ws, i + 2, len(headers), is_alt=(i % 2 == 1))

    # 格式化列宽
    col_widths = [22, 18, 10, 14, 12, 16, 12, 14, 12, 14, 14, 12, 14]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # 设置对齐
    for row in ws.iter_rows(min_row=2, max_row=6, min_col=1, max_col=13):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    freeze_top_row(ws)
    return ws

# ==================== 工作表5: 競爭格局 ====================
def create_competitive_landscape_sheet(wb):
    """创建竞争格局表"""
    ws = wb.create_sheet(title='競爭格局')

    headers = ['廠商', '總部', '營收規模', '核心產品', '技術路線', '目標市場', '主要優勢', '主要劣勢', '評級']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    comp_data = [
        ['Power Integrations', '美國', '$580M', 'InnoSwitch4-CZ', 'GaN/SiC合封', '高端工業/筆電', '最高可靠性', '價格高/迭代慢', 'A+'],
        ['Navitas Semiconductor', '美國', '$180M', 'GaNFast NV6169', 'GaN power IC', '消費旗艦', 'GaN整合最強', '產品線窄', 'B+'],
        ['Infineon Technologies', '德國', '€5.2B', 'EiceDRIVER/CoolGaN', 'GaN+SiC雙平台', '汽車/工業', '寬禁帶最全', '機構複雜', 'A'],
        ['Texas Instruments', '美國', '$17B(總)', 'UCC29950', '數字功率', '工業/醫療/汽車', '模擬+數位整合', '保守定價', 'A'],
        ['南芯半導體', '中國', '私有', 'SC9712/SC8108', '協議+DC-DC SoC', '消費/ODM', '性價比最優', 'GaN落後', 'A-'],
        ['智融科技', '中國', '私有', 'SW3536/SW3516', '多口協議IC', '消費/ODM', '多口市場領導', 'IP薄', 'A-'],
        ['英集芯', '中國', '私有', 'IP2189/IP5356', 'USB PD SoC', '白牌/電商', '最低成本', '質量差異', 'B+'],
        ['东科半导体', '中國', '私有', 'DK8607/DK003', 'ACF控制器', '消費/ODM', '中國最佳ACF', 'GaN路線圖有限', 'B'],
        ['杰华特', '中國', '私有', 'JW7726/JW1510', '數字LLC/PFC', '工業/汽車', '模擬IP強', '消費存在感弱', 'B+'],
        ['矽力杰', '中國', '私有', 'SY6270/SY6801', '高效DC-DC', '工業/計算/消費', '產品線廣', 'GaN晚入局', 'B'],
    ]

    for i, row_data in enumerate(comp_data):
        ws.append(row_data)
        style_data_row(ws, i + 2, len(headers), is_alt=(i % 2 == 1))

    # 格式化列宽
    col_widths = [25, 10, 15, 18, 16, 16, 14, 16, 8]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    freeze_top_row(ws)
    return ws

# ==================== 工作表6: 終端應用 ====================
def create_application_sheet(wb):
    """创建终端应用表"""
    ws = wb.create_sheet(title='終端應用')

    headers = ['應用場景', '功率需求(W)', '關鍵技術要求', '市場規模(億美元)', '2025-2030 CAGR(%)', '主要終端廠商', '晶片需求']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    app_data = [
        ['AI PC/筆電', '140-240', '超薄/多口/96%+效率', 45, 12, '蘋果/戴爾/聯想/華為', 'PD3.1 EPR/GaN共封'],
        ['旗艦手機/摺疊機', '100-150', '迷你化/UFCS+PD雙協議', 38, 8, '蘋果/三星/小米/OPPO', '小型BGA/高效低載'],
        ['遊戲設備', '45-240', '高穩定性/低噪聲', 12, 25, '任天堂/Valve/華碩/索尼', '高瞬態響應'],
        ['AI終端', '100-240', '智能功率分配/多設備', 5, 45, 'Rabbit/ Humane/ Meta', 'VDM/多設備協商'],
        ['工業控制', '65-240', '寬電壓/-40°C~85°C/長壽命', 8, 10, '西門子/施耐德/ABB', '工業級/長生命周期'],
        ['醫療設備', '65-150', 'IEC60601/雙重絕緣/2xMOPP', 5, 14, '飛利普/GE/西門子醫療', '醫療隔離/患者安全'],
        ['商用顯示器', '65-240', 'DP Alt Mode/高分支持', 12, 18, '戴爾/LG/明基', 'USB-C retimer整合'],
        ['電動車800V', '200-240', 'AEC-Q100/ISO26262/雙向', 15, 38, '比亞迪/CATL/特斯拉/蔚來', '車規GaN/雙向PFC'],
        ['戶外儲能', '100-240', '太陽能MPPT/AC輸出/抗震', 8, 30, 'Jackery/EcoFlow/Bluetti', '寬Vin/高效率偏載'],
    ]

    for i, row_data in enumerate(app_data):
        ws.append(row_data)
        style_data_row(ws, i + 2, len(headers), is_alt=(i % 2 == 1))

    # 格式化列宽
    col_widths = [16, 14, 28, 18, 18, 28, 22]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    freeze_top_row(ws)
    return ws

# ==================== 工作表7: 成本毛利 ====================
def create_cost_margin_sheet(wb):
    """创建成本毛利表"""
    ws = wb.create_sheet(title='成本毛利')

    headers = ['功率等級', '低端單價(美元)', '中端單價(美元)', '高端單價(美元)', '市場均價(美元)', '晶片BOM成本(美元)', '晶片廠商毛利(%)', 'ODM毛利(%)', '零售毛利(%)']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    cost_data = [
        ['65W', 3, 6, 12, 6.2, '0.8-1.5', '50-60', '20-30', '35-45'],
        ['100W', 6, 12, 25, 13.5, '1.5-2.5', '48-58', '22-32', '33-43'],
        ['140W', 10, 18, 38, 19.8, '2.5-4.0', '45-55', '25-35', '30-40'],
        ['240W', 18, 30, 55, 32.5, '4.0-7.0', '42-52', '20-30', '28-38'],
    ]

    for i, row_data in enumerate(cost_data):
        ws.append(row_data)
        style_data_row(ws, i + 2, len(headers), is_alt=(i % 2 == 1))

    # 格式化列宽
    col_widths = [14, 16, 16, 16, 16, 18, 16, 14, 14]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # 数字格式
    for row in range(2, 6):
        for col in [2, 3, 4, 5]:
            ws.cell(row=row, column=col).number_format = '$#,##0.00'

    freeze_top_row(ws)
    return ws

# ==================== 工作表8: 技術路線圖 ====================
def create_tech_roadmap_sheet(wb):
    """创建技术路线图表"""
    ws = wb.create_sheet(title='技術路線圖')

    headers = ['年份', '集成水平', '開關頻率(kHz)', '功率密度(W/in³)', '協議標準', '關鍵材料', '主要芯片進展']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    roadmap_data = [
        [2025, 'PFC+LLC dual-chip', '200-500', '20-25', 'PD3.1+UFCS兴起', 'GaN分立+SiC', 'PI InnoSwitch4, Navitas NV61xx'],
        [2026, 'GaN co-pack主流', '500-1000', '25-32', 'PD3.1+UFCS必备', 'GaN共封裝', 'GaN Power IC量產'],
        [2027, 'Single GaN SoC', '1000-2000', '32-40', '多协议融合', 'GaN SoC', 'Single-chip GaN'],
        [2028, 'Full digital single chip', '2000-3000', '40-50', 'AI感知功率分配', 'Digital GaN', 'AI-Power Control'],
        [2029, 'GaN+SiC hybrid', '3000-5000', '50-60', '预测性功率管理', 'GaN+SiC混合', 'Hybrid功率器件'],
        [2030, 'Monolithic GaN SoC', '5000-10000', '60-80', '传感+控制全集成', 'Monolithic GaN', 'Full-integration SoC'],
    ]

    for i, row_data in enumerate(roadmap_data):
        ws.append(row_data)
        style_data_row(ws, i + 2, len(headers), is_alt=(i % 2 == 1))

    # 格式化列宽
    col_widths = [10, 22, 18, 18, 22, 18, 30]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    freeze_top_row(ws)
    return ws

# ==================== 主程序 ====================
def main():
    """主程序入口"""
    print("=" * 60)
    print("PD 大電流適配器集成技術市場分析 - Excel數據文件生成")
    print("=" * 60)

    # 创建工作簿
    wb = Workbook()

    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 创建各工作表
    print("\n[1/8] 创建市场概览表...")
    create_market_overview_sheet(wb)

    print("[2/8] 创建功率结构表...")
    create_power_structure_sheet(wb)

    print("[3/8] 创建区域市场表...")
    create_regional_market_sheet(wb)

    print("[4/8] 创建技术方案对比表...")
    create_tech_comparison_sheet(wb)

    print("[5/8] 创建竞争格局表...")
    create_competitive_landscape_sheet(wb)

    print("[6/8] 创建终端应用表...")
    create_application_sheet(wb)

    print("[7/8] 创建成本毛利表...")
    create_cost_margin_sheet(wb)

    print("[8/8] 创建技术路线图表...")
    create_tech_roadmap_sheet(wb)

    # 保存文件
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pd_market_data.xlsx')
    wb.save(output_path)

    print("\n" + "=" * 60)
    print(f"Excel文件已生成: {output_path}")
    print("=" * 60)
    print("\n工作表列表:")
    for i, sheet in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet}")

if __name__ == "__main__":
    main()
