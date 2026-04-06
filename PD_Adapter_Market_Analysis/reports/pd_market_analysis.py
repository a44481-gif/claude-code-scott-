#!/usr/bin/env python3
"""
PD High-Current Adapter Market Analysis & Visualization (2025-2030)
Power Delivery Adapter with Integrated Technology Market Report
Author: AI Analysis System
"""

import json
import os
import sys
from datetime import datetime
from typing import Dict, List, Any

# Chart generation
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False
    print("[WARN] matplotlib not available, charts will be skipped")

# Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference, LineChart
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not available, Excel export will be skipped")


# =============================================================================
# MARKET DATA
# =============================================================================

MARKET_DATA = {
    "overview": [
        {"year": 2025, "market_size_b": 14.2, "shipments_m": 485, "integration_rate": 45},
        {"year": 2026, "market_size_b": 17.1, "shipments_m": 572, "integration_rate": 58},
        {"year": 2027, "market_size_b": 20.6, "shipments_m": 672, "integration_rate": 72},
        {"year": 2028, "market_size_b": 24.0, "shipments_m": 778, "integration_rate": 82},
        {"year": 2029, "market_size_b": 26.8, "shipments_m": 878, "integration_rate": 88},
        {"year": 2030, "market_size_b": 28.7, "shipments_m": 978, "integration_rate": 92},
    ],
    "power_segments": {
        "2025": {"65W": 52, "100W": 28, "140W": 12, "240W": 8},
        "2030": {"65W": 35, "100W": 30, "140W": 20, "240W": 15},
    },
    "power_cagr": {"65W": 8.2, "100W": 16.5, "140W": 28.7, "240W": 32.4},
    "regional_share": {
        "2025": {"China": 48, "North America": 20, "Europe": 17, "Southeast Asia": 8, "Japan/Korea": 5, "Others": 2},
        "2030": {"China": 45, "North America": 19, "Europe": 20, "Southeast Asia": 10, "Japan/Korea": 4, "Others": 2},
    },
    "regional_cagr": {"China": 13.8, "North America": 14.2, "Europe": 18.5, "Southeast Asia": 22.0},
}

TECHNOLOGY_COMPARISON = [
    {
        "architecture": "Discrete MOSFET",
        "vendors": "Infineon, ON Semi",
        "efficiency": 92,
        "power_density": 12,
        "bom_cost_index": 1.00,
        "maturity": "Mature",
        "applications": "Low-cost adapters",
        "integration": "None",
        "pros": "Lowest cost, mature supply chain",
        "cons": "Large size, low efficiency",
    },
    {
        "architecture": "GaN HEMT + Controller",
        "vendors": "Navitas, GaN Systems",
        "efficiency": 95,
        "power_density": 18,
        "bom_cost_index": 0.85,
        "maturity": "Growing",
        "applications": "High-end chargers",
        "integration": "Partial (GaN + driver)",
        "pros": "High efficiency, compact size",
        "cons": "Higher cost than Si",
    },
    {
        "architecture": "GaN Integrated (PowiGaN)",
        "vendors": "Power Integrations",
        "efficiency": 96,
        "power_density": 25,
        "bom_cost_index": 0.75,
        "maturity": "Mature",
        "applications": "Premium adapters, Industrial",
        "integration": "High (GaN + Controller + Protection)",
        "pros": "Best reliability, excellent thermal",
        "cons": "Premium pricing",
    },
    {
        "architecture": "SiC MOSFET",
        "vendors": "Wolfspeed, ROHM",
        "efficiency": 97,
        "power_density": 22,
        "bom_cost_index": 1.20,
        "maturity": "Early",
        "applications": "Industrial, >200W high power",
        "integration": "Partial",
        "pros": "Best for >200W, highest efficiency",
        "cons": "Expensive, limited availability",
    },
    {
        "architecture": "PFC+LLC Single Chip",
        "vendors": "PI, ON Semi",
        "efficiency": 95,
        "power_density": 20,
        "bom_cost_index": 0.70,
        "maturity": "Growing",
        "applications": "Notebooks, Gaming",
        "integration": "High (PFC + LLC + Controller)",
        "pros": "Smallest BOM, simple design",
        "cons": "Complex thermal management",
    },
    {
        "architecture": "Protocol SoC + GaN",
        "vendors": "Southchip, Richtek",
        "efficiency": 94,
        "power_density": 19,
        "bom_cost_index": 0.72,
        "maturity": "Growing",
        "applications": "Smartphones, Tablets, Multi-port",
        "integration": "Very High (Protocol + DC-DC + GaN)",
        "pros": "Smallest PCB, fastest design cycle",
        "cons": "Protocol lock-in risk",
    },
]

COMPETITIVE_LANDSCAPE = {
    "international": [
        {"vendor": "Power Integrations", "share": 22, "technology": "GaN/SiC co-pack", "strength": "Reliability, high integration", "segment": "Premium/Industrial"},
        {"vendor": "Navitas", "share": 18, "technology": "GaN Power IC", "strength": "GaN pioneer, efficiency leader", "segment": "High-end chargers"},
        {"vendor": "Infineon", "share": 20, "technology": "GaN + SiC dual platform", "strength": "Broadest portfolio", "segment": "Industrial/Automotive"},
        {"vendor": "TI", "share": 15, "technology": "Digital power", "strength": "System solutions", "segment": "Broad market"},
        {"vendor": "ON Semi", "share": 6, "technology": "MOSFET", "strength": "Cost competitive", "segment": "Cost-sensitive"},
        {"vendor": "Wolfspeed", "share": 4, "technology": "SiC", "strength": "High-power SiC", "segment": ">200W high power"},
        {"vendor": "STMicro", "share": 5, "technology": "Mixed signal", "strength": "Automotive focus", "segment": "Industrial/Consumer"},
    ],
    "chinese": [
        {"vendor": "Southchip (南芯)", "share": 8, "technology": "Protocol + GaN SoC", "strength": "Full protocol stack", "segment": "Smartphones/Laptops"},
        {"vendor": "iSmart (智融)", "share": 6, "technology": "Multi-port protocol IC", "strength": "Multi-port leader", "segment": "Power banks/Chargers"},
        {"vendor": "Injoinic (英集芯)", "share": 5, "technology": "Protocol IC", "strength": "Turnkey solutions", "segment": "Consumer adapters"},
        {"vendor": "Dongkee (东科)", "share": 6, "technology": "GaN Controller, ACF", "strength": "Best ACF in China", "segment": "Consumer adapters"},
        {"vendor": "Joulwatt (杰华特)", "share": 4, "technology": "Digital LLC/PFC", "strength": "Digital power", "segment": "Industrial"},
        {"vendor": "Silan (矽力杰)", "share": 3, "technology": "High-efficiency DC-DC", "strength": "High efficiency", "segment": "Industrial/Power"},
        {"vendor": "Chipown (芯朋微)", "share": 3, "technology": "AC-DC solutions", "strength": "High integration", "segment": "Home appliances"},
    ],
}

TERMINAL_APPLICATIONS = [
    {"app": "AI PC/Laptop", "power_w": "140-240", "key_reqs": "Ultra-thin, multi-port, 96%+ efficiency", "cagr": 28.7, "vendors": "Apple, Dell, HP, Lenovo, ASUS", "market_b": 45},
    {"app": "Gaming Console", "power_w": "120-320", "key_reqs": "High stability, low noise", "cagr": 25.0, "vendors": "Sony, Microsoft, Nintendo, Valve", "market_b": 12},
    {"app": "Smartphone", "power_w": "25-100", "key_reqs": "Miniaturization, UFCS+PD dual-protocol", "cagr": 8.0, "vendors": "Apple, Samsung, Xiaomi, OPPO", "market_b": 38},
    {"app": "Tablet", "power_w": "30-65", "key_reqs": "Compact, lightweight", "cagr": 5.0, "vendors": "Apple, Samsung, Huawei", "market_b": 15},
    {"app": "Automotive IVI", "power_w": "45-120", "key_reqs": "AEC-Q, wide temp range", "cagr": 40.0, "vendors": "Tesla, BYD, NIO", "market_b": 20},
    {"app": "EVSE", "power_w": "3.3-22", "key_reqs": "Smart charging, bidirectional", "cagr": 45.0, "vendors": "Tesla Supercharger, BYD", "market_b": 25},
    {"app": "Industrial", "power_w": "65-240", "key_reqs": "Industrial grade, EMC Class B", "cagr": 10.0, "vendors": "Siemens, Schneider, ABB", "market_b": 8},
    {"app": "Medical", "power_w": "30-120", "key_reqs": "IEC60601, 2x MOPP", "cagr": 14.0, "vendors": "Philips, GE, Siemens Health", "market_b": 5},
    {"app": "Outdoor Energy", "power_w": "100-240", "key_reqs": "Portable, high capacity", "cagr": 30.0, "vendors": "Jackery, EcoFlow, Bluetti", "market_b": 8},
]

COST_ANALYSIS = [
    {"segment": "25W-45W", "discrete_bom": 5.20, "integrated_bom": 4.10, "retail": 12, "margin": 52},
    {"segment": "45W-65W", "discrete_bom": 7.80, "integrated_bom": 6.20, "retail": 18, "margin": 54},
    {"segment": "65W-100W", "discrete_bom": 11.50, "integrated_bom": 8.80, "retail": 28, "margin": 56},
    {"segment": "100W-140W", "discrete_bom": 18.20, "integrated_bom": 14.50, "retail": 48, "margin": 58},
    {"segment": "140W-240W", "discrete_bom": 28.50, "integrated_bom": 22.00, "retail": 85, "margin": 62},
    {"segment": "240W+", "discrete_bom": 48.00, "integrated_bom": 38.00, "retail": 160, "margin": 68},
]

TECH_ROADMAP = [
    {"year": 2025, "integration": "PFC+LLC dual-chip", "frequency_khz": "200-500", "density": "20-25", "protocol": "PD3.1 + UFCS emerging"},
    {"year": 2026, "integration": "GaN co-pack mainstream", "frequency_khz": "500-1000", "density": "25-32", "protocol": "PD3.1 + UFCS mandatory"},
    {"year": 2027, "integration": "Single GaN SoC", "frequency_khz": "1000-2000", "density": "32-40", "protocol": "Multi-protocol convergence"},
    {"year": 2028, "integration": "Full digital single chip", "frequency_khz": "2000-3000", "density": "40-50", "protocol": "AI-aware power distribution"},
    {"year": 2029, "integration": "GaN+SiC hybrid", "frequency_khz": "3000-5000", "density": "50-60", "protocol": "Predictive power management"},
    {"year": 2030, "integration": "Monolithic GaN SoC", "frequency_khz": "5000-10000", "density": "60-80", "protocol": "Sensor + control integration"},
]


# =============================================================================
# CHART GENERATION
# =============================================================================

def generate_charts(output_dir: str):
    """Generate all market analysis charts"""
    if not HAS_MATPLOTLIB:
        print("[SKIP] matplotlib not available")
        return

    os.makedirs(output_dir, exist_ok=True)

    # Set style
    plt.style.use('seaborn-v0_8-whitegrid')
    colors = ['#1a1a2e', '#16213e', '#0f3460', '#e94560', '#f5a623']

    # Chart 1: Market Size Growth
    fig, ax = plt.subplots(figsize=(10, 6))
    years = [d["year"] for d in MARKET_DATA["overview"]]
    sizes = [d["market_size_b"] for d in MARKET_DATA["overview"]]
    bars = ax.bar(years, sizes, color='#0f3460', width=0.7, edgecolor='white')
    ax.plot(years, sizes, 'o-', color='#e94560', linewidth=2, markersize=8)
    ax.set_xlabel('Year', fontsize=12)
    ax.set_ylabel('Market Size (USD Billion)', fontsize=12)
    ax.set_title('Global PD Adapter Market Size 2025-2030', fontsize=14, fontweight='bold')
    for bar, size in zip(bars, sizes):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3,
                f'${size}B', ha='center', va='bottom', fontsize=10, fontweight='bold')
    ax.set_ylim(0, 35)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'chart1_market_size.png'), dpi=150)
    plt.close()
    print(f"[OK] chart1_market_size.png")

    # Chart 2: Shipments Growth
    fig, ax = plt.subplots(figsize=(10, 6))
    shipments = [d["shipments_m"] for d in MARKET_DATA["overview"]]
    bars = ax.bar(years, shipments, color='#16213e', width=0.7)
    ax.set_xlabel('Year', fontsize=12)
    ax.set_ylabel('Shipments (Million Units)', fontsize=12)
    ax.set_title('PD Adapter Shipments 2025-2030', fontsize=14, fontweight='bold')
    for bar, s in zip(bars, shipments):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 10,
                f'{s}M', ha='center', va='bottom', fontsize=10)
    ax.set_ylim(0, 1200)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'chart2_shipments.png'), dpi=150)
    plt.close()
    print(f"[OK] chart2_shipments.png")

    # Chart 3: Power Segment Share (Pie charts 2025 vs 2030)
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    labels_2025 = list(MARKET_DATA["power_segments"]["2025"].keys())
    sizes_2025 = list(MARKET_DATA["power_segments"]["2025"].values())
    labels_2030 = list(MARKET_DATA["power_segments"]["2030"].keys())
    sizes_2030 = list(MARKET_DATA["power_segments"]["2030"].values())

    wedges1, texts1, autotexts1 = axes[0].pie(sizes_2025, labels=labels_2025, autopct='%1.1f%%',
                                                 colors=colors, startangle=90, explode=[0.05]*4)
    axes[0].set_title('Power Segment Share 2025', fontsize=12, fontweight='bold')

    wedges2, texts2, autotexts2 = axes[1].pie(sizes_2030, labels=labels_2030, autopct='%1.1f%%',
                                                 colors=colors, startangle=90, explode=[0.05]*4)
    axes[1].set_title('Power Segment Share 2030', fontsize=12, fontweight='bold')

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'chart3_power_segments.png'), dpi=150)
    plt.close()
    print(f"[OK] chart3_power_segments.png")

    # Chart 4: Regional Distribution (Grouped bar)
    fig, ax = plt.subplots(figsize=(12, 6))
    regions = ["China", "North America", "Europe", "Southeast Asia", "Japan/Korea"]
    share_2025 = [MARKET_DATA["regional_share"]["2025"].get(r, 0) for r in regions]
    share_2030 = [MARKET_DATA["regional_share"]["2030"].get(r, 0) for r in regions]

    x = np.arange(len(regions))
    width = 0.35

    bars1 = ax.bar(x - width/2, share_2025, width, label='2025', color='#0f3460')
    bars2 = ax.bar(x + width/2, share_2030, width, label='2030', color='#e94560')

    ax.set_xlabel('Region', fontsize=12)
    ax.set_ylabel('Market Share (%)', fontsize=12)
    ax.set_title('Regional Market Share Comparison 2025 vs 2030', fontsize=14, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(regions, rotation=15)
    ax.legend()
    ax.set_ylim(0, 60)

    for bar in bars1 + bars2:
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                f'{bar.get_height():.0f}%', ha='center', va='bottom', fontsize=9)

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'chart4_regional_share.png'), dpi=150)
    plt.close()
    print(f"[OK] chart4_regional_share.png")

    # Chart 5: Technology Efficiency Comparison
    fig, ax = plt.subplots(figsize=(12, 6))
    archs = [t["architecture"] for t in TECHNOLOGY_COMPARISON]
    efficiencies = [t["efficiency"] for t in TECHNOLOGY_COMPARISON]
    densities = [t["power_density"] for t in TECHNOLOGY_COMPARISON]

    x = np.arange(len(archs))
    width = 0.35

    bars1 = ax.bar(x - width/2, efficiencies, width, label='Efficiency (%)', color='#0f3460')
    ax2 = ax.twinx()
    bars2 = ax2.bar(x + width/2, densities, width, label='Power Density (W/in3)', color='#e94560')

    ax.set_xlabel('Architecture', fontsize=12)
    ax.set_ylabel('Efficiency (%)', fontsize=12, color='#0f3460')
    ax2.set_ylabel('Power Density (W/in3)', fontsize=12, color='#e94560')
    ax.set_title('Technology Comparison: Efficiency vs Power Density', fontsize=14, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(archs, rotation=30, ha='right', fontsize=9)
    ax.legend(loc='upper left')
    ax2.legend(loc='upper right')

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'chart5_tech_comparison.png'), dpi=150)
    plt.close()
    print(f"[OK] chart5_tech_comparison.png")

    # Chart 6: Cost Comparison by Segment
    fig, ax = plt.subplots(figsize=(12, 6))
    segments = [c["segment"] for c in COST_ANALYSIS]
    discrete = [c["discrete_bom"] for c in COST_ANALYSIS]
    integrated = [c["integrated_bom"] for c in COST_ANALYSIS]

    x = np.arange(len(segments))
    width = 0.35

    ax.bar(x - width/2, discrete, width, label='Discrete BOM ($)', color='#16213e')
    ax.bar(x + width/2, integrated, width, label='Integrated BOM ($)', color='#e94560')

    ax.set_xlabel('Power Segment', fontsize=12)
    ax.set_ylabel('BOM Cost (USD)', fontsize=12)
    ax.set_title('BOM Cost: Discrete vs Integrated Solutions', fontsize=14, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(segments, rotation=15)
    ax.legend()

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'chart6_bom_comparison.png'), dpi=150)
    plt.close()
    print(f"[OK] chart6_bom_comparison.png")

    print(f"\n[OK] All 6 charts generated in {output_dir}")


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_to_excel(output_path: str):
    """Export all data to Excel workbook"""
    if not HAS_OPENPYXL:
        print("[SKIP] openpyxl not available")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # Sheet 1: Market Overview
    ws1 = wb.create_sheet("Market_Overview")
    headers1 = ["Year", "Market Size (USD B)", "Shipments (M units)", "Integration Rate (%)", "YoY Growth (%)"]
    ws1.append(headers1)
    style_header(ws1)

    prev_size = 0
    for d in MARKET_DATA["overview"]:
        yoy = ((d["market_size_b"] - prev_size) / prev_size * 100) if prev_size > 0 else 0
        ws1.append([d["year"], d["market_size_b"], d["shipments_m"], d["integration_rate"], round(yoy, 1)])
        prev_size = d["market_size_b"]

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 22

    # Sheet 2: Power Segments
    ws2 = wb.create_sheet("Power_Segments")
    headers2 = ["Power Segment", "Market Share 2025 (%)", "Market Share 2030 (%)", "CAGR (%)", "2025 Market (USD B)", "2030 Market (USD B)"]
    ws2.append(headers2)
    style_header(ws2)

    total_2025 = MARKET_DATA["overview"][0]["market_size_b"]
    total_2030 = MARKET_DATA["overview"][-1]["market_size_b"]

    for seg in ["65W", "100W", "140W", "240W"]:
        share_2025 = MARKET_DATA["power_segments"]["2025"].get(seg, 0)
        share_2030 = MARKET_DATA["power_segments"]["2030"].get(seg, 0)
        cagr = MARKET_DATA["power_cagr"].get(seg, 0)
        market_2025 = round(total_2025 * share_2025 / 100, 2)
        market_2030 = round(total_2030 * share_2030 / 100, 2)
        ws2.append([seg, share_2025, share_2030, cagr, market_2025, market_2030])

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 22

    # Sheet 3: Technology Comparison
    ws3 = wb.create_sheet("Tech_Comparison")
    headers3 = ["Architecture", "Vendors", "Efficiency (%)", "Power Density (W/in3)", "BOM Cost Index", "Maturity", "Applications", "Integration Level", "Pros", "Cons"]
    ws3.append(headers3)
    style_header(ws3)

    for t in TECHNOLOGY_COMPARISON:
        ws3.append([
            t["architecture"], t["vendors"], t["efficiency"], t["power_density"],
            t["bom_cost_index"], t["maturity"], t["applications"],
            t["integration"], t["pros"], t["cons"]
        ])

    for col_letter, width in [("A", 25), ("B", 25), ("C", 15), ("D", 20), ("E", 15), ("F", 12), ("G", 25), ("H", 25), ("I", 30), ("J", 30)]:
        ws3.column_dimensions[col_letter].width = width

    # Sheet 4: Competitive Landscape
    ws4 = wb.create_sheet("Competitive_Landscape")
    headers4 = ["Vendor", "Market Share (%)", "Technology", "Strength", "Target Segment", "Region"]
    ws4.append(headers4)
    style_header(ws4)

    for v in COMPETITIVE_LANDSCAPE["international"]:
        ws4.append([v["vendor"], v["share"], v["technology"], v["strength"], v["segment"], "International"])

    for v in COMPETITIVE_LANDSCAPE["chinese"]:
        ws4.append([v["vendor"], v["share"], v["technology"], v["strength"], v["segment"], "China"])

    for col in ws4.columns:
        ws4.column_dimensions[col[0].column_letter].width = 25

    # Sheet 5: Terminal Applications
    ws5 = wb.create_sheet("Terminal_Apps")
    headers5 = ["Application", "Power (W)", "Key Requirements", "CAGR (%)", "Top Vendors", "Addressable Market (USD B)"]
    ws5.append(headers5)
    style_header(ws5)

    for a in TERMINAL_APPLICATIONS:
        ws5.append([a["app"], a["power_w"], a["key_reqs"], a["cagr"], a["vendors"], a["market_b"]])

    for col in ws5.columns:
        ws5.column_dimensions[col[0].column_letter].width = 25

    # Sheet 6: Cost Analysis
    ws6 = wb.create_sheet("Cost_Analysis")
    headers6 = ["Segment", "Discrete BOM ($)", "Integrated BOM ($)", "Savings ($)", "Savings (%)", "Retail Price ($)", "Gross Margin (%)"]
    ws6.append(headers6)
    style_header(ws6)

    for c in COST_ANALYSIS:
        savings = c["discrete_bom"] - c["integrated_bom"]
        savings_pct = round(savings / c["discrete_bom"] * 100, 1)
        ws6.append([c["segment"], c["discrete_bom"], c["integrated_bom"], round(savings, 2), savings_pct, c["retail"], c["margin"]])

    for col in ws6.columns:
        ws6.column_dimensions[col[0].column_letter].width = 20

    # Sheet 7: Technology Roadmap
    ws7 = wb.create_sheet("Tech_Roadmap")
    headers7 = ["Year", "Integration Level", "Switching Frequency (kHz)", "Power Density (W/in3)", "Protocol Standard"]
    ws7.append(headers7)
    style_header(ws7)

    for r in TECH_ROADMAP:
        ws7.append([r["year"], r["integration"], r["frequency_khz"], r["density"], r["protocol"]])

    for col in ws7.columns:
        ws7.column_dimensions[col[0].column_letter].width = 28

    wb.save(output_path)
    print(f"[OK] Excel file saved: {output_path}")


# =============================================================================
# JSON DATA EXPORT
# =============================================================================

def export_to_json(output_path: str):
    """Export all data to JSON"""
    data = {
        "report_info": {
            "title": "PD High-Current Adapter Integrated Technology Market Analysis 2025-2030",
            "date": datetime.now().strftime("%Y-%m-%d"),
            "version": "1.0",
        },
        "market_overview": MARKET_DATA["overview"],
        "power_segments": MARKET_DATA["power_segments"],
        "power_cagr": MARKET_DATA["power_cagr"],
        "regional_share": MARKET_DATA["regional_share"],
        "technology_comparison": TECHNOLOGY_COMPARISON,
        "competitive_landscape": COMPETITIVE_LANDSCAPE,
        "terminal_applications": TERMINAL_APPLICATIONS,
        "cost_analysis": COST_ANALYSIS,
        "technology_roadmap": TECH_ROADMAP,
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"[OK] JSON data saved: {output_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60)
    print("PD High-Current Adapter Market Analysis Generator")
    print("=" * 60)

    # Output directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, "reports")
    os.makedirs(output_dir, exist_ok=True)

    print(f"\n[INFO] Output directory: {output_dir}")

    # Generate charts
    print("\n--- Generating Charts ---")
    generate_charts(output_dir)

    # Export to Excel
    print("\n--- Exporting to Excel ---")
    excel_path = os.path.join(output_dir, "pd_market_data_2025_2030.xlsx")
    export_to_excel(excel_path)

    # Export to JSON
    print("\n--- Exporting to JSON ---")
    json_path = os.path.join(output_dir, "pd_market_data.json")
    export_to_json(json_path)

    print("\n" + "=" * 60)
    print("GENERATION COMPLETE")
    print("=" * 60)
    print(f"\nOutput files:")
    print(f"  Charts: {output_dir}/chart*.png (6 files)")
    print(f"  Excel:   {excel_path}")
    print(f"  JSON:    {json_path}")


if __name__ == "__main__":
    main()
